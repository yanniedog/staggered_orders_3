#!/usr/bin/env python3
"""
Create Excel Staggered Ladder Calculator - Redesigned

Generates an Excel workbook with formulas that replicate the Python calculator functionality.
Improved layout with better organization, clearer sections, and enhanced usability.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule, CellIsRule
from openpyxl.worksheet.protection import SheetProtection
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def create_helpers_sheet(ws_helpers, ws_calc, input_start_row, buy_start_row, sell_start_row,
                         budget_ref, current_price_ref, profit_target_ref, num_rungs_ref,
                         buy_price_range_ref, helper_font, border):
    """Create Helpers sheet with all helper calculations and protect it."""
    
    # Title
    ws_helpers['A1'] = "HELPER CALCULATIONS"
    ws_helpers['A1'].font = Font(bold=True, size=14)
    ws_helpers['A1'].fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    ws_helpers.merge_cells('A1:B1')
    ws_helpers.row_dimensions[1].height = 25
    
    # Warning message
    ws_helpers['A2'] = "⚠ WARNING: These cells are protected. Modifying them may cause calculation errors."
    ws_helpers['A2'].font = Font(size=10, italic=True, color="FF0000", bold=True)
    ws_helpers['A2'].fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
    ws_helpers.merge_cells('A2:B2')
    ws_helpers.row_dimensions[2].height = 20
    
    # Convert references to include sheet name
    calc_sheet_name = ws_calc.title
    budget_ref_full = f"'{calc_sheet_name}'!{budget_ref}"
    current_price_ref_full = f"'{calc_sheet_name}'!{current_price_ref}"
    profit_target_ref_full = f"'{calc_sheet_name}'!{profit_target_ref}"
    num_rungs_ref_full = f"'{calc_sheet_name}'!{num_rungs_ref}"
    buy_price_range_ref_full = f"'{calc_sheet_name}'!{buy_price_range_ref}"
    
    row = 4
    
    # Section: Buy Price Helpers
    ws_helpers[f'A{row}'] = "Buy Price Helpers"
    ws_helpers[f'A{row}'].font = Font(bold=True, size=11)
    ws_helpers[f'A{row}'].fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    ws_helpers.merge_cells(f'A{row}:B{row}')
    row += 1
    
    # Buy Upper (row 1)
    buy_upper_row = row
    ws_helpers[f'A{row}'] = "Buy Upper"
    ws_helpers[f'A{row}'].font = helper_font
    ws_helpers[f'A{row}'].alignment = Alignment(horizontal='right', vertical='center')
    ws_helpers[f'B{row}'] = f'={current_price_ref_full}'
    ws_helpers[f'B{row}'].number_format = '#,##0.00'
    ws_helpers[f'B{row}'].border = border
    buy_upper_ref = f"'Helpers'!$B${row}"
    row += 1
    
    # Buy Lower (row 2)
    buy_lower_row = row
    ws_helpers[f'A{row}'] = "Buy Lower"
    ws_helpers[f'A{row}'].font = helper_font
    ws_helpers[f'A{row}'].alignment = Alignment(horizontal='right', vertical='center')
    ws_helpers[f'B{row}'] = f'=IF({current_price_ref_full}>0,{current_price_ref_full}*(1-{buy_price_range_ref_full}/100),0)'
    ws_helpers[f'B{row}'].number_format = '#,##0.00'
    ws_helpers[f'B{row}'].border = border
    buy_lower_ref = f"'Helpers'!$B${row}"
    row += 1
    
    # Price Step (row 3)
    price_step_row = row
    ws_helpers[f'A{row}'] = "Price Step"
    ws_helpers[f'A{row}'].font = helper_font
    ws_helpers[f'A{row}'].alignment = Alignment(horizontal='right', vertical='center')
    ws_helpers[f'B{row}'] = f'=IF(AND({num_rungs_ref_full}>1,B{buy_upper_row}>0,B{buy_lower_row}>0),(B{buy_upper_row}-B{buy_lower_row})/({num_rungs_ref_full}-1),IF({current_price_ref_full}>0,{current_price_ref_full}*0.1,0))'
    ws_helpers[f'B{row}'].number_format = '#,##0.00'
    ws_helpers[f'B{row}'].border = border
    price_step_ref = f"'Helpers'!$B${row}"
    row += 1
    
    # Max Buy Price (row 4)
    max_buy_price_row = row
    ws_helpers[f'A{row}'] = "Max Buy Price"
    ws_helpers[f'A{row}'].font = helper_font
    ws_helpers[f'A{row}'].alignment = Alignment(horizontal='right', vertical='center')
    ws_helpers[f'B{row}'] = f'={current_price_ref_full}'
    ws_helpers[f'B{row}'].number_format = '#,##0.00'
    ws_helpers[f'B{row}'].border = border
    max_buy_price_ref = f"'Helpers'!$B${row}"
    row += 1
    
    # Gap Size (row 5)
    gap_size_row = row
    ws_helpers[f'A{row}'] = "Gap Size"
    ws_helpers[f'A{row}'].font = helper_font
    ws_helpers[f'A{row}'].alignment = Alignment(horizontal='right', vertical='center')
    ws_helpers[f'B{row}'] = f'=IF(AND(B{max_buy_price_row}>0,B{price_step_row}>0),MAX(1,CEILING(B{max_buy_price_row}*{profit_target_ref_full}/100*0.5/B{price_step_row},1))*B{price_step_row},0)'
    ws_helpers[f'B{row}'].number_format = '#,##0.00'
    ws_helpers[f'B{row}'].border = border
    gap_size_ref = f"'Helpers'!$B${row}"
    row += 1
    
    # Min Sell Price (row 6)
    min_sell_price_row = row
    ws_helpers[f'A{row}'] = "Min Sell Price"
    ws_helpers[f'A{row}'].font = helper_font
    ws_helpers[f'A{row}'].alignment = Alignment(horizontal='right', vertical='center')
    ws_helpers[f'B{row}'] = f'=IF(AND(B{max_buy_price_row}>0,B{gap_size_row}>0),B{max_buy_price_row}+B{gap_size_row},0)'
    ws_helpers[f'B{row}'].number_format = '#,##0.00'
    ws_helpers[f'B{row}'].border = border
    min_sell_price_ref = f"'Helpers'!$B${row}"
    row += 2
    
    # Section: Buy Order Totals
    ws_helpers[f'A{row}'] = "Buy Order Totals"
    ws_helpers[f'A{row}'].font = Font(bold=True, size=11)
    ws_helpers[f'A{row}'].fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    ws_helpers.merge_cells(f'A{row}:B{row}')
    row += 1
    
    # Total Buy Qty
    total_buy_qty_row = row
    ws_helpers[f'A{row}'] = "Total Buy Qty"
    ws_helpers[f'A{row}'].font = helper_font
    ws_helpers[f'A{row}'].alignment = Alignment(horizontal='right', vertical='center')
    ws_helpers[f'B{row}'] = f"=SUM('{calc_sheet_name}'!G{buy_start_row}:G{buy_start_row+19})"
    ws_helpers[f'B{row}'].number_format = '#,##0.0000'
    ws_helpers[f'B{row}'].border = border
    total_buy_qty_ref = f"'Helpers'!$B${row}"
    row += 1
    
    # Total Cost
    total_cost_row = row
    ws_helpers[f'A{row}'] = "Total Cost"
    ws_helpers[f'A{row}'].font = helper_font
    ws_helpers[f'A{row}'].alignment = Alignment(horizontal='right', vertical='center')
    ws_helpers[f'B{row}'] = f"=SUM('{calc_sheet_name}'!F{buy_start_row}:F{buy_start_row+19})"
    ws_helpers[f'B{row}'].number_format = '#,##0.00'
    ws_helpers[f'B{row}'].border = border
    total_cost_ref = f"'Helpers'!$B${row}"
    row += 1
    
    # Target Revenue
    target_revenue_row = row
    ws_helpers[f'A{row}'] = "Target Revenue"
    ws_helpers[f'A{row}'].font = helper_font
    ws_helpers[f'A{row}'].alignment = Alignment(horizontal='right', vertical='center')
    ws_helpers[f'B{row}'] = f'=B{total_cost_row}*(1+{profit_target_ref_full}/100)'
    ws_helpers[f'B{row}'].number_format = '#,##0.00'
    ws_helpers[f'B{row}'].border = border
    target_revenue_ref = f"'Helpers'!$B${row}"
    row += 2
    
    # Section: Sell Order Helpers
    ws_helpers[f'A{row}'] = "Sell Order Helpers"
    ws_helpers[f'A{row}'].font = Font(bold=True, size=11)
    ws_helpers[f'A{row}'].fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
    ws_helpers.merge_cells(f'A{row}:B{row}')
    row += 1
    
    # Initial Revenue
    initial_revenue_row = row
    ws_helpers[f'A{row}'] = "Initial Revenue"
    ws_helpers[f'A{row}'].font = helper_font
    ws_helpers[f'A{row}'].alignment = Alignment(horizontal='right', vertical='center')
    ws_helpers[f'B{row}'] = f"=SUMPRODUCT('{calc_sheet_name}'!B{sell_start_row}:B{sell_start_row+19},'{calc_sheet_name}'!G{sell_start_row}:G{sell_start_row+19})"
    ws_helpers[f'B{row}'].number_format = '#,##0.00'
    ws_helpers[f'B{row}'].border = border
    initial_revenue_ref = f"'Helpers'!$B${row}"
    row += 1
    
    # Price Scale
    price_scale_row = row
    ws_helpers[f'A{row}'] = "Price Scale"
    ws_helpers[f'A{row}'].font = helper_font
    ws_helpers[f'A{row}'].alignment = Alignment(horizontal='right', vertical='center')
    ws_helpers[f'B{row}'] = f'=IF(AND(B{initial_revenue_row}>0,{target_revenue_ref}>0),{target_revenue_ref}/B{initial_revenue_row},1)'
    ws_helpers[f'B{row}'].number_format = '#,##0.0000'
    ws_helpers[f'B{row}'].border = border
    price_scale_ref = f"'Helpers'!$B${row}"
    row += 1
    
    # Min Valid Sell
    min_valid_sell_row = row
    ws_helpers[f'A{row}'] = "Min Valid Sell"
    ws_helpers[f'A{row}'].font = helper_font
    ws_helpers[f'A{row}'].alignment = Alignment(horizontal='right', vertical='center')
    ws_helpers[f'B{row}'] = f'=IF(AND({max_buy_price_ref}>0,{price_step_ref}>0),{max_buy_price_ref}+{price_step_ref},0)'
    ws_helpers[f'B{row}'].number_format = '#,##0.00'
    ws_helpers[f'B{row}'].border = border
    min_valid_sell_ref = f"'Helpers'!$B${row}"
    row += 2
    
    # Section: Validation Checks
    validation_start_row = row
    ws_helpers[f'A{row}'] = "Validation Checks"
    ws_helpers[f'A{row}'].font = Font(bold=True, size=11)
    ws_helpers[f'A{row}'].fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    ws_helpers.merge_cells(f'A{row}:B{row}')
    row += 1
    
    # Validation formulas - convert references to include sheet name
    calc_sheet_name = ws_calc.title
    validations = [
        ('Budget > 0', f"=IF('{calc_sheet_name}'!{budget_ref}>0,\"✓ OK\",\"✗ ERROR: Budget must be positive\")"),
        ('Current Price > 0', f"=IF('{calc_sheet_name}'!{current_price_ref}>0,\"✓ OK\",\"✗ ERROR: Price must be positive\")"),
        ('Profit Target Valid', f"=IF(AND('{calc_sheet_name}'!{profit_target_ref}>=10,'{calc_sheet_name}'!{profit_target_ref}<=200),\"✓ OK\",\"✗ ERROR: Profit target must be 10-200%\")"),
        ('Num Rungs >= 1', f"=IF('{calc_sheet_name}'!{num_rungs_ref}>=1,\"✓ OK\",\"✗ ERROR: Number of rungs must be at least 1\")"),
        ('Buy Price Range Valid', f"=IF(AND('{calc_sheet_name}'!{buy_price_range_ref}>0,'{calc_sheet_name}'!{buy_price_range_ref}<=100),\"✓ OK\",\"✗ ERROR: Buy price range must be 0-100%\")"),
        ('Buy Lower > 0', f"=IF({buy_lower_ref}>0,\"✓ OK\",\"✗ ERROR: Buy lower price is non-positive. Reduce Buy Price Range %\")"),
        ('Sell Prices > Buy Prices', f"=IF(AND('{calc_sheet_name}'!{num_rungs_ref}>0,'{calc_sheet_name}'!{num_rungs_ref}<=20,SUM('{calc_sheet_name}'!B{buy_start_row}:B{buy_start_row+19})>0,SUM('{calc_sheet_name}'!C{sell_start_row}:C{sell_start_row+19})>0),IF(AND(ISNUMBER(INDEX('{calc_sheet_name}'!C{sell_start_row}:C{sell_start_row+19},1)),ISNUMBER(INDEX('{calc_sheet_name}'!B{buy_start_row}:B{buy_start_row+19},'{calc_sheet_name}'!{num_rungs_ref}))),IF(INDEX('{calc_sheet_name}'!C{sell_start_row}:C{sell_start_row+19},1)>INDEX('{calc_sheet_name}'!B{buy_start_row}:B{buy_start_row+19},'{calc_sheet_name}'!{num_rungs_ref}),\"✓ OK\",\"✗ ERROR: Sell prices must exceed buy prices\"),\"N/A\"),\"N/A\")"),
        ('Volume Match', f"=IF(ABS({total_buy_qty_ref}-SUM('{calc_sheet_name}'!G{sell_start_row}:G{sell_start_row+19}))<0.0001,\"✓ OK\",\"⚠ WARNING: Buy and sell quantities do not match exactly\")"),
    ]
    
    for label, formula in validations:
        ws_helpers[f'A{row}'] = label
        ws_helpers[f'A{row}'].font = Font(size=9)
        ws_helpers[f'A{row}'].alignment = Alignment(horizontal='right', vertical='center')
        ws_helpers[f'B{row}'] = formula
        ws_helpers[f'B{row}'].font = Font(size=9)
        # Color code: green for OK, red for ERROR, orange for WARNING
        if 'ERROR' in formula:
            ws_helpers[f'B{row}'].font = Font(size=9, color="FF0000", bold=True)
            ws_helpers[f'B{row}'].fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
        elif 'WARNING' in formula:
            ws_helpers[f'B{row}'].font = Font(size=9, color="FF8800", bold=True)
            ws_helpers[f'B{row}'].fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
        else:
            ws_helpers[f'B{row}'].font = Font(size=9, color="008800", bold=True)
            ws_helpers[f'B{row}'].fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        ws_helpers[f'B{row}'].alignment = Alignment(horizontal='left', vertical='center')
        ws_helpers[f'B{row}'].border = border
        ws_helpers.row_dimensions[row].height = 16
        row += 1
    
    # Set column widths
    ws_helpers.column_dimensions['A'].width = 20
    ws_helpers.column_dimensions['B'].width = 30
    
    # Protect all cells with formulas (column B) and labels (column A)
    for row_num in range(1, row + 1):
        # Protect formula cells
        cell_b = ws_helpers[f'B{row_num}']
        if cell_b.value:
            cell_b.protection = Protection(locked=True)
        # Protect label cells
        cell_a = ws_helpers[f'A{row_num}']
        cell_a.protection = Protection(locked=True)
    
    # Enable sheet protection with password (optional - can be removed)
    # This will show a warning when user tries to edit protected cells
    ws_helpers.protection.sheet = True
    ws_helpers.protection.formatCells = False  # Allow formatting
    ws_helpers.protection.formatColumns = False
    ws_helpers.protection.formatRows = False
    ws_helpers.protection.insertColumns = False
    ws_helpers.protection.insertRows = False
    ws_helpers.protection.insertHyperlinks = False
    ws_helpers.protection.deleteColumns = False
    ws_helpers.protection.deleteRows = False
    ws_helpers.protection.selectLockedCells = True  # Allow selection
    ws_helpers.protection.selectUnlockedCells = True
    ws_helpers.protection.sort = False
    ws_helpers.protection.autoFilter = False
    ws_helpers.protection.pivotTables = False
    ws_helpers.protection.objects = False
    ws_helpers.protection.scenarios = False
    
    return {
        'buy_upper_ref': buy_upper_ref,
        'buy_lower_ref': buy_lower_ref,
        'price_step_ref': price_step_ref,
        'max_buy_price_ref': max_buy_price_ref,
        'gap_size_ref': gap_size_ref,
        'min_sell_price_ref': min_sell_price_ref,
        'total_buy_qty_ref': total_buy_qty_ref,
        'total_cost_ref': total_cost_ref,
        'target_revenue_ref': target_revenue_ref,
        'initial_revenue_ref': initial_revenue_ref,
        'price_scale_ref': price_scale_ref,
        'min_valid_sell_ref': min_valid_sell_ref
    }

def create_excel_calculator(filename: str = "staggered_ladder_calculator.xlsx"):
    """Create Excel workbook with formulas for staggered ladder calculator - redesigned."""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Calculator"
    
    # Define enhanced styles
    title_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    input_fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
    summary_fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")
    buy_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    sell_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
    
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=18, color="FFFFFF")
    section_font = Font(bold=True, size=13)
    label_font = Font(bold=True, size=10)
    input_font = Font(size=11)
    helper_font = Font(italic=True, size=9, color="666666")
    
    thick_border = Border(
        left=Side(style='thick', color="000000"),
        right=Side(style='thick', color="000000"),
        top=Side(style='thick', color="000000"),
        bottom=Side(style='thick', color="000000")
    )
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ========== TITLE SECTION ==========
    ws['A1'] = "STAGGERED ORDER LADDER CALCULATOR"
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:D1')
    ws.row_dimensions[1].height = 30
    
    # ========== INPUT PARAMETERS SECTION ==========
    row = 3
    ws[f'A{row}'] = "INPUT PARAMETERS"
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].fill = input_fill
    ws.merge_cells(f'A{row}:D{row}')
    ws.row_dimensions[row].height = 22
    row += 1
    
    # Input labels and cells with default values
    input_data = [
        ('Total Budget ($)', 10000, '#,##0.00', 'Total amount available for buy orders'),
        ('Current Price ($)', 50, '#,##0.00', 'Current market price (highest buy rung)'),
        ('Profit Target (%)', 75, '0.00', 'Target profit percentage (10-200%)'),
        ('Number of Rungs', 10, '0', 'Number of buy/sell orders to place'),
        ('Buy Price Range (%)', 30, '0.00', 'Price range for buy ladder (0-100%)')
    ]
    
    # Extract default num_rungs for row hiding
    default_num_rungs = input_data[3][1]  # Get default value from 'Number of Rungs'
    
    input_start_row = row
    for i, (label, default_val, num_format, hint) in enumerate(input_data):
        # Label
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = label_font
        ws[f'A{row}'].alignment = Alignment(horizontal='right', vertical='center')
        ws[f'A{row}'].fill = input_fill
        
        # Input cell
        ws[f'B{row}'] = default_val
        ws[f'B{row}'].number_format = num_format
        ws[f'B{row}'].border = thick_border
        ws[f'B{row}'].font = input_font
        ws[f'B{row}'].fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        ws[f'B{row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Hint/description
        ws[f'C{row}'] = hint
        ws[f'C{row}'].font = Font(size=9, italic=True, color="666666")
        ws[f'C{row}'].alignment = Alignment(horizontal='left', vertical='center')
        
        ws.row_dimensions[row].height = 20
        row += 1
    
    # Store cell references for use in formulas
    budget_ref = f'$B${input_start_row}'
    current_price_ref = f'$B${input_start_row+1}'
    profit_target_ref = f'$B${input_start_row+2}'
    num_rungs_ref = f'$B${input_start_row+3}'
    buy_price_range_ref = f'$B${input_start_row+4}'
    
    row += 1
    
    # ========== SUMMARY RESULTS SECTION ==========
    ws[f'A{row}'] = "SUMMARY RESULTS"
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].fill = summary_fill
    ws.merge_cells(f'A{row}:D{row}')
    ws.row_dimensions[row].height = 22
    row += 1
    
    summary_start_row = row
    
    # Calculate buy_start_row position (after summary section + spacing)
    # Summary has 8 rows (title + 8 data rows), plus 2 rows spacing = 10 rows
    buy_start_row = summary_start_row + 8 + 2  # Summary rows + spacing
    sell_start_row = buy_start_row + 25  # Where sell orders table starts
    
    summary_formulas = [
        ('Total Budget', f'={budget_ref}', '#,##0.00'),
        ('Number of Rungs', f'={num_rungs_ref}', '0'),
        ('Total Cost', f'=SUM(F{buy_start_row}:F{buy_start_row+19})', '#,##0.00'),
        ('Total Revenue', f'=SUM(H{sell_start_row}:H{sell_start_row+19})', '#,##0.00'),
        ('Total Profit', f'=B{summary_start_row+3}-B{summary_start_row+2}', '#,##0.00'),
        ('Profit Percentage', f'=IF(B{summary_start_row+2}>0,(B{summary_start_row+4}/B{summary_start_row+2})*100,0)', '0.00"%"'),
        ('Total Buy Quantity', f'=SUM(G{buy_start_row}:G{buy_start_row+19})', '#,##0.0000'),
        ('Total Sell Quantity', f'=SUM(G{sell_start_row}:G{sell_start_row+19})', '#,##0.0000'),
    ]
    
    for i, (label, formula, num_format) in enumerate(summary_formulas):
        # Label
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = label_font
        ws[f'A{row}'].alignment = Alignment(horizontal='right', vertical='center')
        ws[f'A{row}'].fill = summary_fill
        
        # Value
        ws[f'B{row}'] = formula
        ws[f'B{row}'].number_format = num_format
        ws[f'B{row}'].border = border
        ws[f'B{row}'].font = Font(size=11, bold=True)
        ws[f'B{row}'].fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        ws[f'B{row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Highlight profit row
        if 'Profit' in label and 'Percentage' not in label:
            ws[f'B{row}'].fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
        elif 'Profit Percentage' in label:
            ws[f'B{row}'].fill = PatternFill(start_color="A5D6A7", end_color="A5D6A7", fill_type="solid")
        
        ws.row_dimensions[row].height = 18
        row += 1
    
    row += 2  # Add small spacing after summary
    
    # Create Helpers sheet and move all helper calculations there
    ws_helpers = wb.create_sheet("Helpers")
    helper_refs = create_helpers_sheet(ws_helpers, ws, input_start_row, buy_start_row, sell_start_row, 
                                       budget_ref, current_price_ref, profit_target_ref, num_rungs_ref, 
                                       buy_price_range_ref, helper_font, border)
    
    # Extract references from helper_refs dictionary
    buy_upper_ref = helper_refs['buy_upper_ref']
    buy_lower_ref = helper_refs['buy_lower_ref']
    price_step_ref = helper_refs['price_step_ref']
    max_buy_price_ref = helper_refs['max_buy_price_ref']
    gap_size_ref = helper_refs['gap_size_ref']
    min_sell_price_ref = helper_refs['min_sell_price_ref']
    total_buy_qty_ref = helper_refs['total_buy_qty_ref']
    total_cost_ref = helper_refs['total_cost_ref']
    target_revenue_ref = helper_refs['target_revenue_ref']
    initial_revenue_ref = helper_refs['initial_revenue_ref']
    price_scale_ref = helper_refs['price_scale_ref']
    min_valid_sell_ref = helper_refs['min_valid_sell_ref']
    
    # ========== BUY ORDERS SECTION ==========
    # Start buy orders right after summary (no validation section needed - it's in Helpers sheet)
    row = buy_start_row - 1
    ws[f'A{row}'] = "BUY ORDERS"
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].fill = buy_fill
    ws.merge_cells(f'A{row}:J{row}')
    ws.row_dimensions[row].height = 22
    row += 1
    
    # Buy orders headers (simplified - hide helper columns)
    buy_headers = ['Rung', 'Buy Price', 'Quantity', 'Cost', 'Cumulative Cost', 'Cumulative Qty', 'Avg Buy Price']
    buy_header_cols = [1, 2, 7, 6, 8, 9, 10]  # Map to actual columns (skipping helper columns)
    
    for col_idx, header in zip(buy_header_cols, buy_headers):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    row += 1
    
    # Buy orders data rows (up to 20 rungs supported)
    for i in range(20):
        rung_num = i + 1
        data_row = buy_start_row + i
        
        # Rung number
        ws.cell(row=data_row, column=1).value = rung_num
        ws.cell(row=data_row, column=1).border = border
        ws.cell(row=data_row, column=1).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=data_row, column=1).fill = buy_fill
        
        # Buy Price (ascending: lowest to highest)
        buy_price_formula = f'=IF({rung_num}<={num_rungs_ref},{buy_lower_ref}+({buy_upper_ref}-{buy_lower_ref})*({rung_num}-1)/(IF({num_rungs_ref}>1,{num_rungs_ref}-1,1)),0)'
        ws.cell(row=data_row, column=2).value = buy_price_formula
        ws.cell(row=data_row, column=2).number_format = '#,##0.00'
        ws.cell(row=data_row, column=2).border = border
        ws.cell(row=data_row, column=2).fill = buy_fill
        
        # Exponential weight for buy (higher weight for lower prices) - column 3 (hidden)
        weight_formula = f'=IF({rung_num}<={num_rungs_ref},EXP(2-2*({rung_num}-1)/(IF({num_rungs_ref}>1,{num_rungs_ref}-1,1))),0)'
        ws.cell(row=data_row, column=3).value = weight_formula
        ws.cell(row=data_row, column=3).number_format = '#,##0.0000'
        
        # Normalized weight - column 4 (hidden)
        norm_weight_formula = f'=IF({rung_num}<={num_rungs_ref},IF(SUM(C{buy_start_row}:C{buy_start_row+19})>0,C{data_row}/SUM(C{buy_start_row}:C{buy_start_row+19}),0),0)'
        ws.cell(row=data_row, column=4).value = norm_weight_formula
        ws.cell(row=data_row, column=4).number_format = '#,##0.0000'
        
        # Cost allocation - column 5 (hidden)
        cost_allocation_formula = f'=IF({rung_num}<={num_rungs_ref},IF(ISNUMBER(D{data_row}),{budget_ref}*D{data_row},0),0)'
        ws.cell(row=data_row, column=5).value = cost_allocation_formula
        ws.cell(row=data_row, column=5).number_format = '#,##0.00'
        
        # Normalized Cost = Cost Allocation * (Budget / SUM(all cost allocations)) - column 6
        normalized_cost_formula = f'=IF({rung_num}<={num_rungs_ref},IF(AND(ISNUMBER(E{data_row}),SUM(E{buy_start_row}:E{buy_start_row+19})>0),E{data_row}*({budget_ref}/SUM(E{buy_start_row}:E{buy_start_row+19})),0),0)'
        ws.cell(row=data_row, column=6).value = normalized_cost_formula
        ws.cell(row=data_row, column=6).number_format = '#,##0.00'
        ws.cell(row=data_row, column=6).border = border
        ws.cell(row=data_row, column=6).fill = buy_fill
        
        # Quantity = Normalized Cost / Buy Price - column 7
        quantity_formula = f'=IF({rung_num}<={num_rungs_ref},IF(B{data_row}>0,F{data_row}/B{data_row},0),0)'
        ws.cell(row=data_row, column=7).value = quantity_formula
        ws.cell(row=data_row, column=7).number_format = '#,##0.0000'
        ws.cell(row=data_row, column=7).border = border
        ws.cell(row=data_row, column=7).fill = buy_fill
        
        # Cumulative Cost - column 8
        if i == 0:
            cum_cost_formula = f'=F{data_row}'
        else:
            cum_cost_formula = f'=IF({rung_num}<={num_rungs_ref},F{data_row}+H{buy_start_row+i-1},0)'
        ws.cell(row=data_row, column=8).value = cum_cost_formula
        ws.cell(row=data_row, column=8).number_format = '#,##0.00'
        ws.cell(row=data_row, column=8).border = border
        ws.cell(row=data_row, column=8).fill = buy_fill
        
        # Cumulative Quantity - column 9
        if i == 0:
            cum_qty_formula = f'=G{data_row}'
        else:
            cum_qty_formula = f'=IF({rung_num}<={num_rungs_ref},G{data_row}+I{buy_start_row+i-1},0)'
        ws.cell(row=data_row, column=9).value = cum_qty_formula
        ws.cell(row=data_row, column=9).number_format = '#,##0.0000'
        ws.cell(row=data_row, column=9).border = border
        ws.cell(row=data_row, column=9).fill = buy_fill
        
        # Avg Buy Price = Cumulative Cost / Cumulative Quantity - column 10
        avg_buy_formula = f'=IF({rung_num}<={num_rungs_ref},IF(I{data_row}>0,H{data_row}/I{data_row},0),0)'
        ws.cell(row=data_row, column=10).value = avg_buy_formula
        ws.cell(row=data_row, column=10).number_format = '#,##0.00'
        ws.cell(row=data_row, column=10).border = border
        ws.cell(row=data_row, column=10).fill = buy_fill
        
        # Hide rows that should not be shown (rung > default num_rungs)
        # Set row height to 0 to effectively hide rows beyond default num_rungs
        if rung_num > default_num_rungs:
            ws.row_dimensions[data_row].height = 0
            ws.row_dimensions[data_row].hidden = True
    
    # Hide helper columns (C, D, E)
    ws.column_dimensions['C'].width = 0.1
    ws.column_dimensions['D'].width = 0.1
    ws.column_dimensions['E'].width = 0.1
    
    # ========== SELL ORDERS SECTION ==========
    row = sell_start_row - 1
    ws[f'A{row}'] = "SELL ORDERS"
    ws[f'A{row}'].font = section_font
    ws[f'A{row}'].fill = sell_fill
    ws.merge_cells(f'A{row}:K{row}')
    ws.row_dimensions[row].height = 22
    row += 1
    
    # Sell orders headers (simplified)
    sell_headers = ['Rung', 'Sell Price', 'Quantity', 'Revenue', 'Cumulative Revenue', 'Cumulative Qty', 'Avg Sell Price']
    sell_header_cols = [1, 3, 7, 8, 9, 10, 11]  # Map to actual columns
    
    for col_idx, header in zip(sell_header_cols, sell_headers):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    row += 1
    
    # Sell orders data rows
    for i in range(20):
        rung_num = i + 1
        data_row = sell_start_row + i
        
        # Rung number
        ws.cell(row=data_row, column=1).value = rung_num
        ws.cell(row=data_row, column=1).border = border
        ws.cell(row=data_row, column=1).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=data_row, column=1).fill = sell_fill
        
        # Initial Sell Price (before scaling) - stored in column 2 (hidden)
        initial_sell_price_formula = f'=IF({rung_num}<={num_rungs_ref},{min_sell_price_ref}+{price_step_ref}*({rung_num}-1),0)'
        ws.cell(row=data_row, column=2).value = initial_sell_price_formula
        ws.cell(row=data_row, column=2).number_format = '#,##0.00'
        
        # Final Sell Price (after scaling and gap check) - column 3
        sell_price_formula = f'=IF({rung_num}<={num_rungs_ref},MAX(B{data_row}*{price_scale_ref},{min_valid_sell_ref}+{price_step_ref}*({rung_num}-1)),0)'
        ws.cell(row=data_row, column=3).value = sell_price_formula
        ws.cell(row=data_row, column=3).number_format = '#,##0.00'
        ws.cell(row=data_row, column=3).border = border
        ws.cell(row=data_row, column=3).fill = sell_fill
        
        # Exponential weight for sell (higher weight for higher prices) - column 4 (hidden)
        sell_weight_formula = f'=IF({rung_num}<={num_rungs_ref},EXP(2*({rung_num}-1)/(IF({num_rungs_ref}>1,{num_rungs_ref}-1,1))),0)'
        ws.cell(row=data_row, column=4).value = sell_weight_formula
        ws.cell(row=data_row, column=4).number_format = '#,##0.0000'
        
        # Normalized sell weight - column 5 (hidden)
        normalized_sell_weight_formula = f'=IF({rung_num}<={num_rungs_ref},IF(SUM(D{sell_start_row}:D{sell_start_row+19})>0,D{data_row}/SUM(D{sell_start_row}:D{sell_start_row+19}),0),0)'
        ws.cell(row=data_row, column=5).value = normalized_sell_weight_formula
        ws.cell(row=data_row, column=5).number_format = '#,##0.0000'
        
        # Sell Quantity = TotalBuyQty * normalized_sell_weight - column 6 (hidden)
        sell_qty_before_norm = f'=IF({rung_num}<={num_rungs_ref},{total_buy_qty_ref}*E{data_row},0)'
        ws.cell(row=data_row, column=6).value = sell_qty_before_norm
        ws.cell(row=data_row, column=6).number_format = '#,##0.0000'
        
        # Normalize quantity to match buy qty exactly - column 7
        normalized_sell_qty_formula = f'=IF({rung_num}<={num_rungs_ref},IF(AND(ISNUMBER(F{data_row}),SUM(F{sell_start_row}:F{sell_start_row+19})>0,{total_buy_qty_ref}>0),F{data_row}*({total_buy_qty_ref}/SUM(F{sell_start_row}:F{sell_start_row+19})),0),0)'
        ws.cell(row=data_row, column=7).value = normalized_sell_qty_formula
        ws.cell(row=data_row, column=7).number_format = '#,##0.0000'
        ws.cell(row=data_row, column=7).border = border
        ws.cell(row=data_row, column=7).fill = sell_fill
        
        # Revenue = Final Sell Price * Normalized Quantity - column 8
        revenue_formula = f'=IF({rung_num}<={num_rungs_ref},C{data_row}*G{data_row},0)'
        ws.cell(row=data_row, column=8).value = revenue_formula
        ws.cell(row=data_row, column=8).number_format = '#,##0.00'
        ws.cell(row=data_row, column=8).border = border
        ws.cell(row=data_row, column=8).fill = sell_fill
        
        # Cumulative Revenue - column 9
        if i == 0:
            cum_rev_formula = f'=H{data_row}'
        else:
            cum_rev_formula = f'=IF({rung_num}<={num_rungs_ref},H{data_row}+I{sell_start_row+i-1},0)'
        ws.cell(row=data_row, column=9).value = cum_rev_formula
        ws.cell(row=data_row, column=9).number_format = '#,##0.00'
        ws.cell(row=data_row, column=9).border = border
        ws.cell(row=data_row, column=9).fill = sell_fill
        
        # Cumulative Quantity - column 10
        if i == 0:
            cum_qty_formula = f'=G{data_row}'
        else:
            cum_qty_formula = f'=IF({rung_num}<={num_rungs_ref},G{data_row}+J{sell_start_row+i-1},0)'
        ws.cell(row=data_row, column=10).value = cum_qty_formula
        ws.cell(row=data_row, column=10).number_format = '#,##0.0000'
        ws.cell(row=data_row, column=10).border = border
        ws.cell(row=data_row, column=10).fill = sell_fill
        
        # Avg Sell Price = Cumulative Revenue / Cumulative Quantity - column 11
        avg_sell_formula = f'=IF({rung_num}<={num_rungs_ref},IF(J{data_row}>0,I{data_row}/J{data_row},0),0)'
        ws.cell(row=data_row, column=11).value = avg_sell_formula
        ws.cell(row=data_row, column=11).number_format = '#,##0.00'
        ws.cell(row=data_row, column=11).border = border
        ws.cell(row=data_row, column=11).fill = sell_fill
        
        # Hide rows that should not be shown (rung > default num_rungs)
        # Set row height to 0 to effectively hide rows beyond default num_rungs
        if rung_num > default_num_rungs:
            ws.row_dimensions[data_row].height = 0
            ws.row_dimensions[data_row].hidden = True
    
    # Hide helper columns (B, D, E, F)
    ws.column_dimensions['B'].width = 0.1  # Initial price
    ws.column_dimensions['D'].width = 0.1  # Weight
    ws.column_dimensions['E'].width = 0.1  # Norm weight
    ws.column_dimensions['F'].width = 0.1  # Qty before norm
    
    # Auto-adjust column widths for visible columns
    visible_cols = ['A', 'B', 'F', 'G', 'H', 'I', 'J', 'K', 'C']  # C is sell price
    for col_letter in visible_cols:
        max_length = 0
        for row_num in range(1, min(200, ws.max_row + 1)):
            cell = ws[f'{col_letter}{row_num}']
            try:
                if cell.value is not None:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
            except:
                pass
        if max_length > 0:
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[col_letter].width = adjusted_width
    
    # Set specific widths for better appearance
    ws.column_dimensions['A'].width = 8  # Rung
    ws.column_dimensions['B'].width = 12  # Buy Price
    ws.column_dimensions['C'].width = 12  # Sell Price
    ws.column_dimensions['F'].width = 14  # Cost/Revenue
    ws.column_dimensions['G'].width = 12  # Quantity
    ws.column_dimensions['H'].width = 16  # Cumulative Cost/Revenue
    ws.column_dimensions['I'].width = 14  # Cumulative Qty
    ws.column_dimensions['J'].width = 14  # Avg Price
    
    # Create Charts sheet
    create_charts_sheet(wb, buy_start_row, sell_start_row, num_rungs_ref)
    
    # Add instructions
    instruction_row = sell_start_row + 25
    ws[f'A{instruction_row}'] = "INSTRUCTIONS"
    ws[f'A{instruction_row}'].font = section_font
    ws.merge_cells(f'A{instruction_row}:D{instruction_row}')
    instruction_row += 1
    
    instructions = [
        "1. Enter your parameters in the INPUT PARAMETERS section at the top.",
        "2. Review the SUMMARY RESULTS to see total cost, revenue, and profit.",
        "3. Check the 'Helpers' sheet for validation checks and detailed calculations.",
        "4. View BUY ORDERS and SELL ORDERS tables below.",
        "5. Only active rungs (based on Number of Rungs) are shown - excess rows are hidden.",
        "6. See the Charts sheet for visualizations.",
        "7. Helper calculations are in the 'Helpers' sheet (protected from modification)."
    ]
    
    for instruction in instructions:
        ws[f'A{instruction_row}'] = instruction
        ws[f'A{instruction_row}'].font = Font(size=9, italic=True, color="666666")
        ws.merge_cells(f'A{instruction_row}:D{instruction_row}')
        instruction_row += 1
    
    wb.save(filename)
    logger.info(f"Excel calculator created: {filename}")

def create_charts_sheet(wb, buy_start_row, sell_start_row, num_rungs_ref):
    """Create charts sheet with visualizations matching Python version - completely redesigned."""
    from openpyxl.chart import LineChart, BarChart, ScatterChart, Reference, Series
    from openpyxl.chart.marker import DataPoint as MarkerDataPoint
    
    ws_charts = wb.create_sheet("Charts")
    ws_calc = wb['Calculator']
    
    # ========== CHART 1: BUY LADDER - Price (Individual and Average) ==========
    chart1 = LineChart()
    chart1.title = "Buy Ladder: Price by Order Rung"
    chart1.style = 13
    chart1.height = 12
    chart1.width = 18
    chart1.legend.position = 'r'
    
    # Categories (rung numbers)
    cats = Reference(ws_calc, min_col=1, min_row=buy_start_row, max_row=buy_start_row+19)
    
    # Individual Buy Price (line with markers)
    data_price = Reference(ws_calc, min_col=2, min_row=buy_start_row-1, max_row=buy_start_row+19)
    series_price = chart1.add_data(data_price, titles_from_data=True)
    try:
        series_price.graphicalProperties.line.solidFill = "1B5E20"  # Dark green
        series_price.graphicalProperties.line.width = 28575  # ~2.5pt
        series_price.marker = MarkerDataPoint()
        series_price.marker.symbol = "circle"
        series_price.marker.size = 9
        series_price.marker.graphicalProperties.solidFill = "1B5E20"
        series_price.marker.graphicalProperties.line.solidFill = "1B5E20"
    except:
        pass  # Fallback if marker styling not supported
    
    # Average Buy Price (dashed line with square markers)
    data_avg = Reference(ws_calc, min_col=10, min_row=buy_start_row-1, max_row=buy_start_row+19)
    series_avg = chart1.add_data(data_avg, titles_from_data=True)
    try:
        series_avg.graphicalProperties.line.solidFill = "2E7D32"  # Medium green
        series_avg.graphicalProperties.line.width = 28575
        series_avg.graphicalProperties.line.dashStyle = "sysDash"
        series_avg.marker = MarkerDataPoint()
        series_avg.marker.symbol = "square"
        series_avg.marker.size = 7
        series_avg.marker.graphicalProperties.solidFill = "2E7D32"
        series_avg.marker.graphicalProperties.line.solidFill = "2E7D32"
    except:
        pass
    
    chart1.set_categories(cats)
    chart1.y_axis.title = 'Price ($)'
    chart1.y_axis.scaling.min = 0
    chart1.x_axis.title = 'Buy Order'
    
    # Position chart
    ws_charts.add_chart(chart1, "A1")
    
    # ========== CHART 1B: BUY LADDER - Quantity ==========
    chart1b = BarChart()
    chart1b.type = "col"
    chart1b.style = 10
    chart1b.title = "Buy Ladder: Quantity by Order Rung"
    chart1b.height = 10
    chart1b.width = 18
    chart1b.legend.position = 'r'
    
    data_qty = Reference(ws_calc, min_col=7, min_row=buy_start_row-1, max_row=buy_start_row+19)
    series_qty = chart1b.add_data(data_qty, titles_from_data=True)
    try:
        series_qty.graphicalProperties.solidFill = "81C784"  # Light green
        series_qty.graphicalProperties.line.solidFill = "2E7D32"  # Dark green border
        series_qty.graphicalProperties.line.width = 11430
    except:
        pass
    
    chart1b.set_categories(cats)
    chart1b.y_axis.title = 'Quantity'
    chart1b.x_axis.title = 'Buy Order'
    chart1b.y_axis.scaling.min = 0
    
    ws_charts.add_chart(chart1b, "A15")
    
    # ========== CHART 2: SELL LADDER - Price (Individual and Average) ==========
    chart2 = LineChart()
    chart2.title = "Sell Ladder: Price by Order Rung"
    chart2.style = 13
    chart2.height = 12
    chart2.width = 18
    chart2.legend.position = 'r'
    
    # Categories (rung numbers)
    cats2 = Reference(ws_calc, min_col=1, min_row=sell_start_row, max_row=sell_start_row+19)
    
    # Individual Sell Price (line with markers)
    data_sell_price = Reference(ws_calc, min_col=3, min_row=sell_start_row-1, max_row=sell_start_row+19)
    series_sell_price = chart2.add_data(data_sell_price, titles_from_data=True)
    try:
        series_sell_price.graphicalProperties.line.solidFill = "B71C1C"  # Dark red
        series_sell_price.graphicalProperties.line.width = 28575
        series_sell_price.marker = MarkerDataPoint()
        series_sell_price.marker.symbol = "circle"
        series_sell_price.marker.size = 9
        series_sell_price.marker.graphicalProperties.solidFill = "B71C1C"
        series_sell_price.marker.graphicalProperties.line.solidFill = "B71C1C"
    except:
        pass
    
    # Average Sell Price (dashed line with square markers)
    data_sell_avg = Reference(ws_calc, min_col=11, min_row=sell_start_row-1, max_row=sell_start_row+19)
    series_sell_avg = chart2.add_data(data_sell_avg, titles_from_data=True)
    try:
        series_sell_avg.graphicalProperties.line.solidFill = "C62828"  # Medium red
        series_sell_avg.graphicalProperties.line.width = 28575
        series_sell_avg.graphicalProperties.line.dashStyle = "sysDash"
        series_sell_avg.marker = MarkerDataPoint()
        series_sell_avg.marker.symbol = "square"
        series_sell_avg.marker.size = 7
        series_sell_avg.marker.graphicalProperties.solidFill = "C62828"
        series_sell_avg.marker.graphicalProperties.line.solidFill = "C62828"
    except:
        pass
    
    chart2.set_categories(cats2)
    chart2.y_axis.title = 'Price ($)'
    chart2.y_axis.scaling.min = 0
    chart2.x_axis.title = 'Sell Order'
    
    # Position chart
    ws_charts.add_chart(chart2, "P1")
    
    # ========== CHART 2B: SELL LADDER - Quantity ==========
    chart2b = BarChart()
    chart2b.type = "col"
    chart2b.style = 10
    chart2b.title = "Sell Ladder: Quantity by Order Rung"
    chart2b.height = 10
    chart2b.width = 18
    chart2b.legend.position = 'r'
    
    data_sell_qty = Reference(ws_calc, min_col=7, min_row=sell_start_row-1, max_row=sell_start_row+19)
    series_sell_qty = chart2b.add_data(data_sell_qty, titles_from_data=True)
    try:
        series_sell_qty.graphicalProperties.solidFill = "EF5350"  # Light red
        series_sell_qty.graphicalProperties.line.solidFill = "C62828"  # Dark red border
        series_sell_qty.graphicalProperties.line.width = 11430
    except:
        pass
    
    chart2b.set_categories(cats2)
    chart2b.y_axis.title = 'Quantity'
    chart2b.x_axis.title = 'Sell Order'
    chart2b.y_axis.scaling.min = 0
    
    ws_charts.add_chart(chart2b, "P15")
    
    # ========== CHART 3: COMBINED BUY/SELL VISUALIZATION ==========
    # Create a scatter chart showing buy and sell orders together
    chart3 = ScatterChart()
    chart3.title = "Combined Buy/Sell Ladder: Average Price vs Order Price"
    chart3.style = 13
    chart3.height = 12
    chart3.width = 20
    chart3.legend.position = 'r'
    
    # For buy orders: X = Buy Price, Y = Avg Buy Price
    buy_price_ref = Reference(ws_calc, min_col=2, min_row=buy_start_row, max_row=buy_start_row+19)
    buy_avg_ref = Reference(ws_calc, min_col=10, min_row=buy_start_row, max_row=buy_start_row+19)
    
    series_buy = Series(values=buy_avg_ref, xvalues=buy_price_ref, title="Buy Orders")
    chart3.series.append(series_buy)
    try:
        series_buy.graphicalProperties.line.solidFill = "2E7D32"
        series_buy.graphicalProperties.line.width = 28575
        series_buy.marker = MarkerDataPoint()
        series_buy.marker.symbol = "circle"
        series_buy.marker.size = 8
        series_buy.marker.graphicalProperties.solidFill = "81C784"
        series_buy.marker.graphicalProperties.line.solidFill = "1B5E20"
        series_buy.marker.graphicalProperties.line.width = 11430
    except:
        pass
    
    # For sell orders: X = Sell Price, Y = Avg Sell Price
    sell_price_ref = Reference(ws_calc, min_col=3, min_row=sell_start_row, max_row=sell_start_row+19)
    sell_avg_ref = Reference(ws_calc, min_col=11, min_row=sell_start_row, max_row=sell_start_row+19)
    
    series_sell = Series(values=sell_avg_ref, xvalues=sell_price_ref, title="Sell Orders")
    chart3.series.append(series_sell)
    try:
        series_sell.graphicalProperties.line.solidFill = "C62828"
        series_sell.graphicalProperties.line.width = 28575
        series_sell.marker = MarkerDataPoint()
        series_sell.marker.symbol = "square"
        series_sell.marker.size = 8
        series_sell.marker.graphicalProperties.solidFill = "EF5350"
        series_sell.marker.graphicalProperties.line.solidFill = "B71C1C"
        series_sell.marker.graphicalProperties.line.width = 11430
    except:
        pass
    
    chart3.x_axis.title = 'Order Price ($)'
    chart3.y_axis.title = 'Average Price ($)'
    chart3.x_axis.scaling.min = 0
    chart3.y_axis.scaling.min = 0
    
    # Position chart
    ws_charts.add_chart(chart3, "A30")
    
    # ========== CHART 4: VOLUME COMPARISON (Buy vs Sell Quantities) ==========
    chart4 = BarChart()
    chart4.type = "col"
    chart4.style = 10
    chart4.title = "Order Quantities: Buy vs Sell Comparison"
    chart4.height = 10
    chart4.width = 18
    chart4.legend.position = 'r'
    
    # Buy quantities
    buy_qty_ref = Reference(ws_calc, min_col=7, min_row=buy_start_row-1, max_row=buy_start_row+19)
    series_buy_qty = chart4.add_data(buy_qty_ref, titles_from_data=True)
    try:
        series_buy_qty.graphicalProperties.solidFill = "81C784"  # Light green
        series_buy_qty.graphicalProperties.line.solidFill = "2E7D32"  # Dark green border
        series_buy_qty.graphicalProperties.line.width = 11430
    except:
        pass
    
    # Sell quantities
    sell_qty_ref = Reference(ws_calc, min_col=7, min_row=sell_start_row-1, max_row=sell_start_row+19)
    series_sell_qty = chart4.add_data(sell_qty_ref, titles_from_data=True)
    try:
        series_sell_qty.graphicalProperties.solidFill = "EF5350"  # Light red
        series_sell_qty.graphicalProperties.line.solidFill = "C62828"  # Dark red border
        series_sell_qty.graphicalProperties.line.width = 11430
    except:
        pass
    
    # Use buy rungs as categories (they should match)
    chart4.set_categories(cats)
    chart4.y_axis.title = 'Quantity'
    chart4.x_axis.title = 'Order Rung'
    chart4.y_axis.scaling.min = 0
    
    # Position chart
    ws_charts.add_chart(chart4, "P30")
    
    logger.info("Charts sheet created with enhanced visualizations")

if __name__ == '__main__':
    import sys
    filename = sys.argv[1] if len(sys.argv) > 1 else "staggered_ladder_calculator.xlsx"
    create_excel_calculator(filename)
    print(f"Excel calculator created successfully: {filename}")
