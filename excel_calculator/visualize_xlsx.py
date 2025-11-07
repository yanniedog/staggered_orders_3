#!/usr/bin/env python3
"""
XLSX File Visualizer - Reads Excel files, extracts charts, and visualizes them
Allows discussion of actual XLSX file contents and chart visualization
"""

from openpyxl import load_workbook
from openpyxl.chart import LineChart, BarChart, ScatterChart, AreaChart, PieChart
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, IconSetRule
import json
import sys
import os

try:
    import matplotlib.pyplot as plt
    import matplotlib.patches as mpatches
    from matplotlib.patches import Rectangle
    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False
    print("Warning: matplotlib not available. Chart visualization will be limited.")

def get_column_letter(col_idx):
    """Convert column index to letter."""
    from openpyxl.utils import get_column_letter
    return get_column_letter(col_idx)

def parse_xlsx_comprehensive(filename):
    """Comprehensively parse an XLSX file and extract all information."""
    print(f"\n{'='*80}")
    print(f"Parsing XLSX File: {filename}")
    print(f"{'='*80}\n")
    
    if not os.path.exists(filename):
        print(f"Error: File '{filename}' not found!")
        return None
    
    wb = load_workbook(filename, data_only=False)  # data_only=False to see formulas
    
    result = {
        'filename': filename,
        'file_size': os.path.getsize(filename),
        'sheets': []
    }
    
    for sheet_name in wb.sheetnames:
        print(f"Processing sheet: {sheet_name}")
        ws = wb[sheet_name]
        
        sheet_info = {
            'name': sheet_name,
            'max_row': ws.max_row,
            'max_column': ws.max_column,
            'merged_cells': [str(mc) for mc in ws.merged_cells.ranges],
            'conditional_formatting': [],
            'charts': [],
            'data_regions': {},
            'formulas': [],
            'styles': {}
        }
        
        # Get conditional formatting rules
        if hasattr(ws, 'conditional_formatting'):
            for range_str, rules in ws.conditional_formatting._cf_rules.items():
                for rule in rules:
                    rule_info = {
                        'range': range_str,
                        'type': type(rule).__name__
                    }
                    if isinstance(rule, ColorScaleRule):
                        rule_info['start_color'] = rule.start_color.rgb if rule.start_color else None
                        rule_info['end_color'] = rule.end_color.rgb if rule.end_color else None
                    elif isinstance(rule, DataBarRule):
                        rule_info['color'] = rule.color.rgb if rule.color else None
                    elif isinstance(rule, IconSetRule):
                        rule_info['icon_style'] = rule.iconSet
                    sheet_info['conditional_formatting'].append(rule_info)
        
        # Extract charts
        if hasattr(ws, '_charts'):
            for idx, chart in enumerate(ws._charts):
                chart_info = extract_chart_info(chart, ws)
                sheet_info['charts'].append(chart_info)
        
        # Extract data from key regions
        sheet_info['data_regions'] = extract_data_regions(ws)
        
        # Extract formulas
        sheet_info['formulas'] = extract_formulas(ws)
        
        # Extract style information
        sheet_info['styles'] = extract_styles(ws)
        
        result['sheets'].append(sheet_info)
    
    return result

def extract_chart_info(chart, worksheet):
    """Extract detailed information about a chart."""
    # Get axis titles safely
    x_axis_title = None
    y_axis_title = None
    if hasattr(chart, 'x_axis') and chart.x_axis:
        if hasattr(chart.x_axis, 'title') and chart.x_axis.title:
            if hasattr(chart.x_axis.title, 'value'):
                x_axis_title = chart.x_axis.title.value
            else:
                x_axis_title = str(chart.x_axis.title)
    if hasattr(chart, 'y_axis') and chart.y_axis:
        if hasattr(chart.y_axis, 'title') and chart.y_axis.title:
            if hasattr(chart.y_axis.title, 'value'):
                y_axis_title = chart.y_axis.title.value
            else:
                y_axis_title = str(chart.y_axis.title)
    
    # Get chart title safely
    chart_title = None
    if hasattr(chart, 'title') and chart.title:
        if hasattr(chart.title, 'value'):
            chart_title = chart.title.value
        else:
            chart_title = str(chart.title)
    
    chart_info = {
        'type': type(chart).__name__,
        'title': chart_title,
        'position': None,
        'series': [],
        'categories': None,
        'x_axis_title': x_axis_title,
        'y_axis_title': y_axis_title,
    }
    
    # Try to get chart position
    if hasattr(chart, 'anchor'):
        chart_info['position'] = str(chart.anchor)
    
    # Extract series data
    if hasattr(chart, 'series'):
        for series in chart.series:
            # Get series title safely
            series_title = None
            if hasattr(series, 'title') and series.title:
                if hasattr(series.title, 'value'):
                    series_title = series.title.value
                elif isinstance(series.title, str):
                    series_title = series.title
                else:
                    series_title = str(series.title)
            
            series_info = {
                'title': series_title,
                'data_range': None,
                'categories_range': None
            }
            
            if hasattr(series, 'values') and series.values:
                series_info['data_range'] = str(series.values)
            
            if hasattr(series, 'categories') and series.categories:
                series_info['categories_range'] = str(series.categories)
            
            chart_info['series'].append(series_info)
    
    return chart_info

def extract_data_regions(ws):
    """Extract data from key regions of the worksheet."""
    regions = {}
    
    # Extract all non-empty cells with their values and formatting
    for row in range(1, min(ws.max_row + 1, 200)):  # Limit to first 200 rows
        for col in range(1, min(ws.max_column + 1, 50)):  # Limit to first 50 columns
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                col_letter = get_column_letter(col)
                cell_ref = f"{col_letter}{row}"
                
                cell_info = {
                    'value': cell.value,
                    'is_formula': isinstance(cell.value, str) and cell.value.startswith('='),
                    'number_format': cell.number_format,
                    'fill_color': None,
                    'font': {
                        'bold': cell.font.bold,
                        'size': cell.font.size,
                        'color': cell.font.color.rgb if cell.font.color and hasattr(cell.font.color, 'rgb') else None,
                        'name': cell.font.name
                    },
                    'alignment': {
                        'horizontal': cell.alignment.horizontal,
                        'vertical': cell.alignment.vertical
                    },
                    'has_border': bool(cell.border.left.style)
                }
                
                # Get fill color
                if cell.fill.start_color:
                    if hasattr(cell.fill.start_color, 'rgb'):
                        cell_info['fill_color'] = cell.fill.start_color.rgb
                    elif hasattr(cell.fill.start_color, 'index'):
                        cell_info['fill_color'] = f"Index: {cell.fill.start_color.index}"
                
                regions[cell_ref] = cell_info
    
    return regions

def extract_formulas(ws):
    """Extract all formulas from the worksheet."""
    formulas = []
    for row in range(1, min(ws.max_row + 1, 200)):
        for col in range(1, min(ws.max_column + 1, 50)):
            cell = ws.cell(row=row, column=col)
            if isinstance(cell.value, str) and cell.value.startswith('='):
                formulas.append({
                    'cell': f"{get_column_letter(col)}{row}",
                    'formula': cell.value
                })
    return formulas

def extract_styles(ws):
    """Extract style information from the worksheet."""
    styles = {
        'column_widths': {},
        'row_heights': {},
        'freeze_panes': ws.freeze_panes
    }
    
    # Column widths
    for col_letter, dim in ws.column_dimensions.items():
        if dim.width:
            styles['column_widths'][col_letter] = dim.width
    
    # Row heights
    for row_num, dim in ws.row_dimensions.items():
        if dim.height:
            styles['row_heights'][row_num] = dim.height
    
    return styles

def visualize_charts_from_xlsx(filename, output_dir="chart_visualizations"):
    """Extract chart data from XLSX and visualize using matplotlib."""
    if not MATPLOTLIB_AVAILABLE:
        print("matplotlib not available. Cannot visualize charts.")
        return
    
    print(f"\n{'='*80}")
    print(f"Visualizing Charts from: {filename}")
    print(f"{'='*80}\n")
    
    # Load with data_only=True to get calculated values, but we need the workbook structure
    wb = load_workbook(filename, data_only=True)
    
    os.makedirs(output_dir, exist_ok=True)
    
    chart_count = 0
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        if hasattr(ws, '_charts') and ws._charts:
            print(f"Found {len(ws._charts)} chart(s) in sheet '{sheet_name}'")
            
            for idx, chart in enumerate(ws._charts):
                chart_count += 1
                try:
                    visualize_chart(chart, wb, sheet_name, idx, output_dir)
                except Exception as e:
                    print(f"  Error visualizing chart {idx+1}: {e}")
                    import traceback
                    traceback.print_exc()
    
    if chart_count == 0:
        print("No charts found in the workbook.")
    else:
        print(f"\nVisualized {chart_count} chart(s). Images saved to '{output_dir}/'")

def visualize_chart(chart, workbook, sheet_name, chart_idx, output_dir):
    """Visualize a single chart from the worksheet."""
    chart_type = type(chart).__name__
    
    # Get chart title
    chart_title = "Untitled Chart"
    if hasattr(chart, 'title') and chart.title:
        if hasattr(chart.title, 'value'):
            chart_title = chart.title.value
        elif hasattr(chart.title, 'tx'):
            if hasattr(chart.title.tx, 'rich'):
                if hasattr(chart.title.tx.rich, 'p'):
                    if chart.title.tx.rich.p:
                        if hasattr(chart.title.tx.rich.p[0], 'r'):
                            if chart.title.tx.rich.p[0].r:
                                if hasattr(chart.title.tx.rich.p[0].r[0], 't'):
                                    chart_title = chart.title.tx.rich.p[0].r[0].t
    
    print(f"  Visualizing {chart_type} chart {chart_idx+1}: {chart_title}")
    
    # Extract data from chart series
    all_series_data = []
    categories_data = None
    
    print(f"      Debug: Chart type: {chart_type}")
    print(f"      Debug: Has series attr: {hasattr(chart, 'series')}")
    
    if hasattr(chart, 'series'):
        print(f"      Debug: Number of series: {len(chart.series)}")
        for idx, series in enumerate(chart.series):
            print(f"      Debug: Series {idx+1}:")
            print(f"        All attributes: {[a for a in dir(series) if not a.startswith('_')]}")
            print(f"        Has values attr: {hasattr(series, 'values')}")
            if hasattr(series, 'values'):
                print(f"        Values is not None: {series.values is not None}")
                print(f"        Values: {series.values}")
                print(f"        Values type: {type(series.values)}")
            
            # Get series title
            series_title = f'Series {len(all_series_data)+1}'
            if hasattr(series, 'title'):
                if hasattr(series.title, 'value'):
                    series_title = series.title.value
                elif isinstance(series.title, str):
                    series_title = series.title
            
            # Extract data values
            data_values = []
            x_values = None  # For scatter charts
            
            # Debug: Print reference info
            if hasattr(series, 'values') and series.values:
                ref_obj = series.values
                print(f"      Debug: Series '{series_title}' values reference:")
                print(f"        Type: {type(ref_obj)}")
                print(f"        Has min_col: {hasattr(ref_obj, 'min_col')}")
                if hasattr(ref_obj, 'min_col'):
                    print(f"        Range: col {ref_obj.min_col}-{ref_obj.max_col}, row {ref_obj.min_row}-{ref_obj.max_row}")
                print(f"        Attributes: {[a for a in dir(ref_obj) if not a.startswith('_')]}")
            
            # Handle different chart types
            if chart_type == 'ScatterChart':
                # ScatterChart uses xvalues and values separately
                if hasattr(series, 'values') and series.values:
                    data_values = extract_reference_values(workbook, series.values)
                if hasattr(series, 'xvalues') and series.xvalues:
                    x_values = extract_reference_values(workbook, series.xvalues)
                
                if data_values and x_values:
                    all_series_data.append({
                        'title': series_title,
                        'data': data_values,
                        'x_values': x_values
                    })
            else:
                # LineChart, BarChart, etc. use values and categories
                if hasattr(series, 'values') and series.values:
                    data_values = extract_reference_values(workbook, series.values)
                
                # Get categories (usually shared across series)
                if categories_data is None and hasattr(series, 'categories') and series.categories:
                    categories_data = extract_reference_values(workbook, series.categories)
                
                if data_values:
                    all_series_data.append({
                        'title': series_title,
                        'data': data_values
                    })
    
    if not all_series_data:
        print(f"    Warning: Could not extract data for chart {chart_idx+1}")
        return
    
    # Create visualization
    fig, ax = plt.subplots(figsize=(12, 8))
    
    # Get axis titles safely
    x_axis_title = 'X Axis'
    y_axis_title = 'Y Axis'
    if hasattr(chart, 'x_axis') and chart.x_axis:
        if hasattr(chart.x_axis, 'title') and chart.x_axis.title:
            if hasattr(chart.x_axis.title, 'value'):
                x_axis_title = chart.x_axis.title.value
    if hasattr(chart, 'y_axis') and chart.y_axis:
        if hasattr(chart.y_axis, 'title') and chart.y_axis.title:
            if hasattr(chart.y_axis.title, 'value'):
                y_axis_title = chart.y_axis.title.value
    
    # Determine chart type and plot accordingly
    if chart_type == 'LineChart':
        for series in all_series_data:
            if categories_data and len(categories_data) == len(series['data']):
                ax.plot(categories_data, series['data'], marker='o', label=series['title'], linewidth=2, markersize=6)
            else:
                ax.plot(series['data'], marker='o', label=series['title'], linewidth=2, markersize=6)
        ax.set_xlabel(x_axis_title)
        ax.set_ylabel(y_axis_title)
        ax.legend()
        ax.grid(True, alpha=0.3)
    
    elif chart_type == 'ScatterChart':
        for series in all_series_data:
            if 'x_values' in series and series['x_values']:
                ax.scatter(series['x_values'], series['data'], label=series['title'], s=100, alpha=0.7)
            else:
                ax.scatter(range(len(series['data'])), series['data'], label=series['title'], s=100, alpha=0.7)
        ax.set_xlabel(x_axis_title)
        ax.set_ylabel(y_axis_title)
        ax.legend()
        ax.grid(True, alpha=0.3)
    
    elif chart_type == 'BarChart':
        x_pos = range(len(all_series_data[0]['data'])) if all_series_data else []
        width = 0.8 / len(all_series_data) if len(all_series_data) > 1 else 0.8
        
        for idx, series in enumerate(all_series_data):
            offset = (idx - len(all_series_data)/2 + 0.5) * width if len(all_series_data) > 1 else 0
            if categories_data and len(categories_data) == len(series['data']):
                ax.bar([x + offset for x in x_pos], series['data'], width, label=series['title'])
                ax.set_xticks(x_pos)
                ax.set_xticklabels([str(c) for c in categories_data], rotation=45, ha='right')
            else:
                ax.bar([x + offset for x in x_pos], series['data'], width, label=series['title'])
        ax.set_xlabel(x_axis_title)
        ax.set_ylabel(y_axis_title)
        ax.legend()
        ax.grid(True, alpha=0.3, axis='y')
    
    elif chart_type == 'AreaChart':
        for series in all_series_data:
            if categories_data and len(categories_data) == len(series['data']):
                ax.fill_between(categories_data, series['data'], alpha=0.6, label=series['title'])
            else:
                ax.fill_between(range(len(series['data'])), series['data'], alpha=0.6, label=series['title'])
        ax.set_xlabel(x_axis_title)
        ax.set_ylabel(y_axis_title)
        ax.legend()
        ax.grid(True, alpha=0.3)
    
    else:
        # Generic fallback
        for series in all_series_data:
            if categories_data and len(categories_data) == len(series['data']):
                ax.plot(categories_data, series['data'], marker='o', label=series['title'])
            else:
                ax.plot(series['data'], marker='o', label=series['title'])
        ax.legend()
        ax.grid(True, alpha=0.3)
    
    # Set title
    ax.set_title(chart_title, fontsize=14, fontweight='bold')
    
    plt.tight_layout()
    
    # Save visualization
    safe_sheet_name = sheet_name.replace('/', '_').replace('\\', '_')
    safe_title = chart_title.replace('/', '_').replace('\\', '_')[:50]
    output_filename = f"{output_dir}/{safe_sheet_name}_chart_{chart_idx+1}_{safe_title}.png"
    plt.savefig(output_filename, dpi=150, bbox_inches='tight')
    plt.close()
    
    print(f"    Saved: {output_filename}")

def extract_reference_values(workbook, ref):
    """Extract values from a Reference object, handling cross-sheet references."""
    try:
        # Handle Reference objects
        if hasattr(ref, 'min_col') and hasattr(ref, 'min_row'):
            # Get the worksheet - Reference objects may have different ways to store sheet info
            ws = None
            
            # Method 1: Check if ref has a worksheet attribute
            if hasattr(ref, 'worksheet') and ref.worksheet:
                ws = ref.worksheet
            # Method 2: Check for sheet name attribute
            elif hasattr(ref, 'sheet') and ref.sheet:
                sheet_name = ref.sheet if isinstance(ref.sheet, str) else str(ref.sheet)
                if sheet_name in workbook.sheetnames:
                    ws = workbook[sheet_name]
            # Method 3: Try to get from data_source (some Reference objects store it here)
            elif hasattr(ref, 'data_source'):
                # This might contain sheet information
                pass
            
            # If still no worksheet, try to find it from the workbook
            # Charts often reference the "Calculator" sheet
            if not ws:
                # Try common sheet names
                for sheet_name in ['Calculator', 'Data', workbook.active.title]:
                    if sheet_name in workbook.sheetnames:
                        ws = workbook[sheet_name]
                        break
            
            if not ws:
                return []
            
            values = []
            # Extract values from the range
            for row in range(ref.min_row, ref.max_row + 1):
                row_values = []
                for col in range(ref.min_col, ref.max_col + 1):
                    cell = ws.cell(row=row, column=col)
                    value = cell.value
                    # Convert to number if possible, otherwise use 0
                    if isinstance(value, (int, float)):
                        row_values.append(value)
                    elif value is None:
                        row_values.append(0)
                    else:
                        try:
                            row_values.append(float(value))
                        except (ValueError, TypeError):
                            row_values.append(0)
                
                # If single column, flatten; if multiple columns, keep as row
                if len(row_values) == 1:
                    values.append(row_values[0])
                else:
                    values.extend(row_values)
            
            # Filter out trailing zeros that are likely empty cells
            # But be careful - actual data might end with zeros
            # For now, return all values
            return values
        
        # Handle string references (like "Calculator!A1:A10")
        elif isinstance(ref, str):
            # Parse sheet!range format
            if '!' in ref:
                sheet_name, range_str = ref.split('!', 1)
                sheet_name = sheet_name.strip("'\"")
                if sheet_name in workbook.sheetnames:
                    ws = workbook[sheet_name]
                    # Simple range parsing (e.g., "A1:A10")
                    # This is simplified - would need more robust parsing
                    return []
            return []
        
    except Exception as e:
        print(f"      Error extracting reference values: {e}")
        import traceback
        traceback.print_exc()
        return []
    
    return []

def print_xlsx_summary(structure):
    """Print a human-readable summary of the XLSX structure."""
    if not structure:
        return
    
    print(f"\n{'='*80}")
    print(f"XLSX FILE SUMMARY: {structure['filename']}")
    print(f"File Size: {structure['file_size']:,} bytes")
    print(f"{'='*80}\n")
    
    for sheet in structure['sheets']:
        print(f"\n[SHEET] {sheet['name']}")
        print(f"   Dimensions: {sheet['max_row']} rows x {sheet['max_column']} columns")
        
        if sheet['merged_cells']:
            print(f"   Merged Cells: {len(sheet['merged_cells'])} ranges")
        
        if sheet['conditional_formatting']:
            print(f"   Conditional Formatting: {len(sheet['conditional_formatting'])} rules")
            for cf in sheet['conditional_formatting']:
                print(f"      - {cf['type']} on {cf['range']}")
        
        if sheet['charts']:
            print(f"   Charts: {len(sheet['charts'])}")
            for idx, chart in enumerate(sheet['charts']):
                print(f"      {idx+1}. {chart['type']}: {chart.get('title', 'Untitled')}")
                print(f"         Series: {len(chart.get('series', []))}")
        
        if sheet['formulas']:
            print(f"   Formulas: {len(sheet['formulas'])}")
        
        if sheet['data_regions']:
            print(f"   Data Cells: {len(sheet['data_regions'])}")
            
            # Show sample of key data
            print(f"\n   Sample Data (first 30 cells with values):")
            count = 0
            for cell_ref, cell_info in sorted(sheet['data_regions'].items()):
                if count >= 30:
                    break
                value_str = str(cell_info['value'])[:40]
                # Remove or replace problematic Unicode characters
                value_str = value_str.encode('ascii', 'replace').decode('ascii')
                if cell_info['is_formula']:
                    value_str = f"={value_str}"
                fill_str = f" [Fill: {cell_info['fill_color']}]" if cell_info['fill_color'] else ""
                bold_str = " [BOLD]" if cell_info['font']['bold'] else ""
                print(f"      {cell_ref}: {value_str}{fill_str}{bold_str}")
                count += 1

def main():
    """Main function."""
    if len(sys.argv) < 2:
        print("Usage: python visualize_xlsx.py <xlsx_file> [--visualize-charts]")
        print("\nExample:")
        print("  python visualize_xlsx.py staggered_ladder_calculator.xlsx")
        print("  python visualize_xlsx.py staggered_ladder_calculator.xlsx --visualize-charts")
        sys.exit(1)
    
    filename = sys.argv[1]
    visualize_charts = '--visualize-charts' in sys.argv
    
    # Parse the file
    structure = parse_xlsx_comprehensive(filename)
    
    if structure:
        # Print summary
        print_xlsx_summary(structure)
        
        # Save JSON structure
        json_filename = filename.replace('.xlsx', '_structure.json')
        with open(json_filename, 'w') as f:
            json.dump(structure, f, indent=2, default=str)
        print(f"\n[FILE] Full structure saved to: {json_filename}")
        
        # Visualize charts if requested
        if visualize_charts:
            visualize_charts_from_xlsx(filename)
        else:
            print("\n[TIP] Use --visualize-charts flag to generate chart images")
    else:
        print("Failed to parse XLSX file.")
        sys.exit(1)

if __name__ == '__main__':
    main()

