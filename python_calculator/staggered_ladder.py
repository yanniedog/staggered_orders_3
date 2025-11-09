#!/usr/bin/env python3
"""
Simplified Staggered Order Ladder Calculator

Calculates buy and sell ladder orders based on user-specified price rungs
or profit targets. Outputs results to Excel and PDF with visualizations.
"""

import pandas as pd
import numpy as np
import logging
from datetime import datetime
from typing import Any, Callable, Dict, List, Tuple, Optional
import sys
import os
import argparse
from dataclasses import dataclass

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)

@dataclass(frozen=True)
class LadderShapeProfile:
    label: str
    description: str
    buy_weight_factory: Callable[[int], np.ndarray]
    sell_weight_factory: Callable[[int], np.ndarray]


def _exp_profile(start: float, stop: float, num: int) -> np.ndarray:
    return np.exp(np.linspace(start, stop, num))


def _linear_profile(start: float, stop: float, num: int) -> np.ndarray:
    return np.linspace(start, stop, num)


LADDER_SHAPE_PROFILES: Dict[str, LadderShapeProfile] = {
    "balanced": LadderShapeProfile(
        label="Balanced",
        description="Even sizing across all rungs.",
        buy_weight_factory=lambda n: np.ones(n, dtype=float),
        sell_weight_factory=lambda n: np.ones(n, dtype=float)
    ),
    "deep-dive": LadderShapeProfile(
        label="Deep Dive",
        description="Heavier budget at deeper buy levels, more supply toward higher sells.",
        buy_weight_factory=lambda n: _exp_profile(1.4, 0.0, n),
        sell_weight_factory=lambda n: _exp_profile(0.0, 1.4, n)
    ),
    "springboard": LadderShapeProfile(
        label="Springboard",
        description="Lighter lower buys, heavier sizing near the top with quick profit taking.",
        buy_weight_factory=lambda n: _exp_profile(0.0, 1.2, n),
        sell_weight_factory=lambda n: _exp_profile(1.2, 0.0, n)
    ),
    "glide-path": LadderShapeProfile(
        label="Glide Path",
        description="Gradual ramp into buys and measured releases on sells.",
        buy_weight_factory=lambda n: _linear_profile(0.6, 1.4, n),
        sell_weight_factory=lambda n: _linear_profile(0.8, 1.2, n)
    )
}


def normalize_ladder_shape_key(value: str) -> str:
    """
    Normalize user input for ladder shape selection to a known profile key.

    Accepts dictionary keys, human-readable labels, or variations with spaces/hyphens.
    """
    key_candidate = value.strip().lower()
    canonical_candidate = key_candidate.replace(" ", "-")

    if canonical_candidate in LADDER_SHAPE_PROFILES:
        return canonical_candidate

    for candidate_key, profile in LADDER_SHAPE_PROFILES.items():
        label_normalized = profile.label.strip().lower()
        label_canonical = label_normalized.replace(" ", "-")
        if key_candidate == label_normalized or canonical_candidate == label_canonical:
            return candidate_key

    raise ValueError(
        f"Invalid ladder shape '{value}'. Choose from: "
        f"{', '.join(sorted(LADDER_SHAPE_PROFILES.keys()))}"
    )


def compute_ladder_depth_pct(buy_prices: List[float]) -> float:
    """
    Compute ladder depth as percentage difference between highest and lowest buy prices.
    """
    if not buy_prices:
        return 0.0
    lowest_price = min(buy_prices)
    highest_price = max(buy_prices)
    if highest_price == 0:
        return 0.0
    return ((highest_price - lowest_price) / highest_price) * 100


def summarize_ladder_shape(results: Dict[str, Any]) -> str:
    """
    Build a human-readable summary of the selected ladder shape.
    """
    label = results.get('ladder_shape_label', "") if isinstance(results, dict) else ""
    description = results.get('ladder_shape_description', "") if isinstance(results, dict) else ""
    if label and description:
        return f"{label} - {description}"
    if label:
        return label
    if description:
        return description
    return "N/A"


class StaggeredLadderCalculator:
    """Calculate staggered order ladders for buy and sell orders."""
    
    def __init__(self, budget: float, num_rungs: int = 10):
        """
        Initialize calculator.
        
        Args:
            budget: Total budget for buy orders
            num_rungs: Number of rungs in the ladder
        """
        # Input validation
        if budget <= 0:
            raise ValueError("Budget must be positive")
        if num_rungs < 1:
            raise ValueError("Number of rungs must be at least 1")
        if not isinstance(num_rungs, int):
            raise ValueError("Number of rungs must be an integer")
        
        self.budget = budget
        self.num_rungs = num_rungs
        self.buy_prices = []
        self.buy_quantities = []
        self.sell_prices = []
        self.sell_quantities = []
        self.ladder_shape_key = ""
        self.ladder_shape_label = ""
    
    def _normalize_budget(self):
        """Normalize buy quantities to exactly match budget."""
        total_cost = sum(qty * price for qty, price in zip(self.buy_quantities, self.buy_prices))
        if total_cost > 0:
            scale_factor = self.budget / total_cost
            self.buy_quantities = [qty * scale_factor for qty in self.buy_quantities]
    
    def _validate_prices(self):
        """Validate that sell prices exceed buy prices for all rungs."""
        # Ensure each sell price exceeds its corresponding buy price
        for i in range(self.num_rungs):
            if self.sell_prices[i] <= self.buy_prices[i]:
                raise ValueError(
                    f"Sell price ${self.sell_prices[i]:.2f} at rung {i+1} must exceed "
                    f"buy price ${self.buy_prices[i]:.2f}"
                )
        
        # Ensure top buy order is below bottom sell order
        max_buy_price = max(self.buy_prices)
        min_sell_price = min(self.sell_prices)
        if max_buy_price >= min_sell_price:
            raise ValueError(
                f"Top buy order (${max_buy_price:.2f}) must be below bottom sell order (${min_sell_price:.2f}). "
                f"Gap needed: ${min_sell_price - max_buy_price:.2f}"
            )
        
    def calculate_from_profit_target(
        self,
        buy_upper: float,
        profit_target: float,
        buy_price_range_pct: float = 30.0,
        ladder_shape_key: str = "deep-dive"
    ) -> Dict:
        """
        Calculate ladder from upper buy rung and profit target using a shape profile.
        Shape profiles control buy/sell quantity weighting across rungs.
        
        Args:
            buy_upper: Upper rung of buy ladder (highest buy price)
            profit_target: Target profit percentage (e.g., 50 for 50%)
            buy_price_range_pct: Percentage range for buy ladder (default 30%)
            ladder_shape_key: Key identifying the ladder shape profile (default "deep-dive")
        
        Returns:
            Dictionary with calculated orders and statistics
        """
        # Input validation
        if buy_upper <= 0:
            raise ValueError("Buy upper must be positive")
        if profit_target < 10 or profit_target > 200:
            raise ValueError("Profit target must be between 10% and 200%")
        if buy_price_range_pct <= 0 or buy_price_range_pct > 100:
            raise ValueError("Buy price range percentage must be between 0 and 100")
        
        shape_key_normalized = ladder_shape_key.strip().lower()
        if shape_key_normalized not in LADDER_SHAPE_PROFILES:
            valid_keys = ', '.join(LADDER_SHAPE_PROFILES.keys())
            raise ValueError(f"Invalid ladder shape '{ladder_shape_key}'. Choose from: {valid_keys}")
        shape_profile = LADDER_SHAPE_PROFILES[shape_key_normalized]
        self.ladder_shape_key = shape_key_normalized
        self.ladder_shape_label = shape_profile.label
        
        profit_multiplier = 1 + (profit_target / 100)
        
        # Overall strategy returns profit_target%
        # Calculate buy prices: ascending order (lowest to highest)
        price_range = buy_upper * (buy_price_range_pct / 100)
        buy_lower = buy_upper - price_range
        if buy_lower <= 0:
            raise ValueError(f"Buy price range results in non-positive buy_lower. "
                           f"Try reducing buy_price_range_pct from {buy_price_range_pct}%")
        
        self.buy_prices = np.linspace(buy_lower, buy_upper, self.num_rungs)
        
        # Calculate consistent price step from buy prices
        if self.num_rungs > 1:
            price_step = (buy_upper - buy_lower) / (self.num_rungs - 1)
        else:
            price_step = buy_upper * 0.1  # Default step if only one rung
        
        # Calculate buy quantities using the selected shape profile
        buy_weights = np.asarray(shape_profile.buy_weight_factory(self.num_rungs), dtype=float)
        if buy_weights.shape[0] != self.num_rungs:
            raise ValueError(
                f"Buy weight factory for shape '{shape_profile.label}' returned {buy_weights.shape[0]} weights,"
                f" expected {self.num_rungs}."
            )
        if not np.isfinite(buy_weights).all():
            raise ValueError(f"Buy weights for shape '{shape_profile.label}' must be finite values")
        if (buy_weights < 0).any():
            raise ValueError(f"Buy weights for shape '{shape_profile.label}' must be non-negative")
        buy_weight_sum = buy_weights.sum()
        if buy_weight_sum <= 0:
            raise ValueError(f"Buy weights for shape '{shape_profile.label}' must sum to a positive value")
        buy_weights = buy_weights / buy_weight_sum
        self.buy_quantities = []
        for weight, price in zip(buy_weights, self.buy_prices):
            if price <= 0:
                raise ValueError("Calculated buy price must be positive")
            self.buy_quantities.append((self.budget * weight) / price)
        
        # Normalize to exactly match budget
        self._normalize_budget()
        
        # Calculate total cost (should equal budget after normalization)
        total_cost = sum(qty * price for qty, price in zip(self.buy_quantities, self.buy_prices))
        
        # Calculate target revenue
        target_revenue = total_cost * profit_multiplier
        
        # Calculate total buy quantity
        total_buy_qty = sum(self.buy_quantities)
        
        # Allocate sell quantities using the selected shape profile
        sell_weights = np.asarray(shape_profile.sell_weight_factory(self.num_rungs), dtype=float)
        if sell_weights.shape[0] != self.num_rungs:
            raise ValueError(
                f"Sell weight factory for shape '{shape_profile.label}' returned {sell_weights.shape[0]} weights,"
                f" expected {self.num_rungs}."
            )
        if not np.isfinite(sell_weights).all():
            raise ValueError(f"Sell weights for shape '{shape_profile.label}' must be finite values")
        if (sell_weights < 0).any():
            raise ValueError(f"Sell weights for shape '{shape_profile.label}' must be non-negative")
        sell_weight_sum = sell_weights.sum()
        if sell_weight_sum <= 0:
            raise ValueError(f"Sell weights for shape '{shape_profile.label}' must sum to a positive value")
        sell_weights = sell_weights / sell_weight_sum
        self.sell_quantities = [total_buy_qty * weight for weight in sell_weights]
        
        # Calculate sell prices with consistent spacing
        # Use the same price_step as buy prices
        max_buy_price = max(self.buy_prices)
        
        # Calculate gap based on profit target to ensure sell prices start high enough
        # For a profit target, we need sell prices to be significantly above buy prices
        # Use profit_target% of max_buy_price as base gap, but ensure at least one price step
        base_gap_pct = profit_target / 100.0  # e.g., 0.5 for 50% profit target
        base_gap_size = max_buy_price * base_gap_pct * 0.5  # Use 50% of profit target as initial gap
        # Round to nearest price_step multiple
        gap_steps = max(1, int(np.ceil(base_gap_size / price_step)))
        gap_size = gap_steps * price_step
        min_sell_price = max_buy_price + gap_size
        
        # Create sell prices with consistent spacing using the same price_step
        self.sell_prices = np.array([min_sell_price + i * price_step for i in range(self.num_rungs)])
        
        # CRITICAL: Ensure sell quantities match buy quantities exactly BEFORE revenue adjustment
        # This ensures volume matching is maintained throughout
        total_sell_qty_initial = sum(self.sell_quantities)
        if total_sell_qty_initial > 0 and abs(total_sell_qty_initial - total_buy_qty) > 0.0001:
            normalize_factor = total_buy_qty / total_sell_qty_initial
            self.sell_quantities = [qty * normalize_factor for qty in self.sell_quantities]
            logger.debug(f"Normalized sell quantities to match buy quantities: {total_buy_qty:.6f}")
        
        # Calculate current revenue with these quantities and prices
        current_revenue = sum(price * qty for price, qty in zip(self.sell_prices, self.sell_quantities))
        
        # Adjust sell PRICES (not quantities) to achieve target revenue while maintaining quantity matching
        # This preserves volume matching while achieving profit target
        if current_revenue > 0 and abs(current_revenue - target_revenue) > 0.01:
            price_scale = target_revenue / current_revenue
            # Scale all sell prices proportionally to achieve target revenue
            # This maintains relative spacing while adjusting absolute levels
            self.sell_prices = self.sell_prices * price_scale
            logger.debug(f"Adjusted sell prices by factor {price_scale:.6f} to achieve target revenue")
            
            # CRITICAL: After scaling, ensure sell prices are still above buy prices
            # If scaling brought prices down too much, shift them up to maintain minimum gap
            max_buy_price_after = max(self.buy_prices)
            min_sell_price_after = min(self.sell_prices)
            required_gap = price_step  # Maintain at least one price step gap
            min_valid_sell_price = max_buy_price_after + required_gap
            
            if min_sell_price_after < min_valid_sell_price:
                # Calculate how much we need to shift prices up
                price_shift = min_valid_sell_price - min_sell_price_after
                self.sell_prices = self.sell_prices + price_shift
                logger.info(f"Shifted sell prices up by ${price_shift:.2f} to maintain gap above buy prices")
                
                # Recalculate revenue with adjusted prices
                adjusted_revenue = sum(price * qty for price, qty in zip(self.sell_prices, self.sell_quantities))
                actual_profit_pct = ((adjusted_revenue - total_cost) / total_cost * 100) if total_cost > 0 else 0
                logger.warning(f"Profit target adjustment: Requested {profit_target:.2f}%, achieved {actual_profit_pct:.2f}% "
                             f"(revenue: ${adjusted_revenue:,.2f} vs target: ${target_revenue:,.2f})")
        
        # CRITICAL: Re-verify quantities still match after any adjustments
        final_total_buy_qty = sum(self.buy_quantities)
        final_total_sell_qty = sum(self.sell_quantities)
        qty_diff = abs(final_total_buy_qty - final_total_sell_qty)
        if qty_diff > 0.0001:
            logger.error(f"QUANTITY MISMATCH DETECTED: buy={final_total_buy_qty:.6f}, sell={final_total_sell_qty:.6f}, diff={qty_diff:.6f}")
            # Force correction: normalize sell quantities to match buy quantities exactly
            if final_total_sell_qty > 0:
                normalize_factor = final_total_buy_qty / final_total_sell_qty
                self.sell_quantities = [qty * normalize_factor for qty in self.sell_quantities]
                final_total_sell_qty_corrected = sum(self.sell_quantities)
                logger.info(f"FORCED CORRECTION: Normalized sell quantities to match buy quantities exactly. "
                          f"New total: {final_total_sell_qty_corrected:.6f} (target: {final_total_buy_qty:.6f})")
            else:
                raise ValueError(f"Cannot normalize sell quantities: total is zero or negative")
        else:
            logger.info(f"Volume matching verified: buy={final_total_buy_qty:.6f}, sell={final_total_sell_qty:.6f}, diff={qty_diff:.9f}")
        
        # Validate that sell prices exceed buy prices and gap exists
        self._validate_prices()
        
        return self._generate_results()
    
    def _generate_results(self) -> Dict:
        """Generate results dictionary with statistics."""
        # Calculate cumulative statistics
        buy_price_count = int(np.size(self.buy_prices))
        sell_price_count = int(np.size(self.sell_prices))
        bottom_buy_price = float(np.min(self.buy_prices)) if buy_price_count else 0.0
        top_sell_price = float(np.max(self.sell_prices)) if sell_price_count else 0.0
        shape_profile = LADDER_SHAPE_PROFILES.get(self.ladder_shape_key)
        shape_label = self.ladder_shape_label or (shape_profile.label if shape_profile else "")
        shape_description = shape_profile.description if shape_profile else ""
        cumulative_buy_cost = []
        cumulative_buy_qty = []
        cumulative_sell_revenue = []
        cumulative_sell_qty = []
        avg_buy_prices = []
        avg_sell_prices = []
        
        # Current price is the highest buy price (baseline)
        current_price = max(self.buy_prices) if len(self.buy_prices) > 0 else 0
        
        # Calculate averages relative to current price
        # For buy orders: orders execute as price drops from highest to lowest
        # buy_prices[0] = lowest price, buy_prices[-1] = highest price (current_price)
        # When price drops to buy_prices[i], orders from buy_prices[-1] down to buy_prices[i] have executed
        # So avg_buy_prices[i] = weighted average of orders from index i to end (inclusive)
        for i in range(self.num_rungs):
            # Include all buy orders from index i to the end (up to current price)
            # This represents: if price drops to buy_prices[i], what orders have executed?
            # Answer: orders at buy_prices[i], buy_prices[i+1], ..., buy_prices[-1]
            total_cost_from_level = 0.0
            total_qty_from_level = 0.0
            
            for j in range(i, self.num_rungs):
                total_cost_from_level += self.buy_quantities[j] * self.buy_prices[j]
                total_qty_from_level += self.buy_quantities[j]
            
            if total_qty_from_level > 0:
                avg_buy_price = total_cost_from_level / total_qty_from_level
            else:
                avg_buy_price = 0.0
            
            cumulative_buy_cost.append(total_cost_from_level)
            cumulative_buy_qty.append(total_qty_from_level)
            
            # Validate average is within expected range (weighted average should be between min and max)
            prices_in_range = [self.buy_prices[j] for j in range(i, self.num_rungs)]
            if prices_in_range:
                min_price_in_range = min(prices_in_range)
                max_price_in_range = max(prices_in_range)
                # Allow small floating point tolerance
                tolerance = 0.01
                if avg_buy_price < min_price_in_range - tolerance or avg_buy_price > max_price_in_range + tolerance:
                    logger.warning(f"Buy average {avg_buy_price:.2f} at index {i} is outside expected price range "
                                 f"[{min_price_in_range:.2f}, {max_price_in_range:.2f}]. "
                                 f"Prices: {prices_in_range}, Quantities: {[self.buy_quantities[j] for j in range(i, self.num_rungs)]}")
            
            avg_buy_prices.append(avg_buy_price)
        
        # For sell orders: at each level, include all orders from lowest sell price UP TO that level
        # sell_prices[0] = lowest price, sell_prices[-1] = highest price
        # When price rises to sell_prices[i], all orders from sell_prices[0] to sell_prices[i] execute
        # So avg_sell_prices[i] = weighted average of orders from index 0 to i (inclusive)
        for i in range(self.num_rungs):
            # Include all sell orders from index 0 to i (from lowest up to current level)
            # This means if price rises to sell_prices[i], all orders at or below this level would execute
            total_revenue_to_level = 0.0
            total_qty_to_level = 0.0
            
            for j in range(i + 1):  # 0 to i (inclusive)
                total_revenue_to_level += self.sell_quantities[j] * self.sell_prices[j]
                total_qty_to_level += self.sell_quantities[j]
            
            if total_qty_to_level > 0:
                avg_sell_price = total_revenue_to_level / total_qty_to_level
            else:
                avg_sell_price = 0.0
            
            # Validate average is within expected range (weighted average should be between min and max)
            prices_in_range = [self.sell_prices[j] for j in range(i + 1)]
            if prices_in_range:
                min_price_in_range = min(prices_in_range)
                max_price_in_range = max(prices_in_range)
                # Allow small floating point tolerance
                tolerance = 0.01
                if avg_sell_price < min_price_in_range - tolerance or avg_sell_price > max_price_in_range + tolerance:
                    logger.warning(f"Sell average {avg_sell_price:.2f} at index {i} is outside expected price range "
                                 f"[{min_price_in_range:.2f}, {max_price_in_range:.2f}]. "
                                 f"Prices: {prices_in_range}, Quantities: {[self.sell_quantities[j] for j in range(i + 1)]}")
            
            avg_sell_prices.append(avg_sell_price)
        
        # Calculate cumulative values (for display purposes)
        running_sell_revenue = 0
        running_sell_qty = 0
        
        for i in range(self.num_rungs):
            running_sell_revenue += self.sell_quantities[i] * self.sell_prices[i]
            running_sell_qty += self.sell_quantities[i]
            cumulative_sell_revenue.append(running_sell_revenue)
            cumulative_sell_qty.append(running_sell_qty)
        
        # CRITICAL: Verify and enforce total buy and sell quantities match exactly
        # This is a final safety check to ensure volume matching is always maintained
        total_buy_qty = sum(self.buy_quantities)
        total_sell_qty = sum(self.sell_quantities)
        qty_diff = abs(total_buy_qty - total_sell_qty)
        qty_tolerance = 0.0001
        
        if qty_diff > qty_tolerance:
            logger.error(f"VOLUME MISMATCH DETECTED IN RESULTS: Total buy quantity ({total_buy_qty:.6f}) != Total sell quantity ({total_sell_qty:.6f}), difference: {qty_diff:.6f}")
            # Force correction by normalizing sell quantities to match buy quantities exactly
            if total_sell_qty > 0:
                normalize_factor = total_buy_qty / total_sell_qty
                self.sell_quantities = [qty * normalize_factor for qty in self.sell_quantities]
                total_sell_qty_corrected = sum(self.sell_quantities)
                qty_diff_after = abs(total_buy_qty - total_sell_qty_corrected)
                logger.info(f"FORCED VOLUME CORRECTION: Normalized sell quantities. "
                          f"New total = {total_sell_qty_corrected:.6f} (target: {total_buy_qty:.6f}), "
                          f"remaining diff: {qty_diff_after:.9f}")
                # Recalculate cumulative sell quantities and revenue after correction
                running_sell_revenue = 0
                running_sell_qty = 0
                cumulative_sell_revenue = []
                cumulative_sell_qty = []
                for i in range(self.num_rungs):
                    running_sell_revenue += self.sell_quantities[i] * self.sell_prices[i]
                    running_sell_qty += self.sell_quantities[i]
                    cumulative_sell_revenue.append(running_sell_revenue)
                    cumulative_sell_qty.append(running_sell_qty)
                # Also recalculate avg_sell_prices since quantities changed
                avg_sell_prices = []
                for i in range(self.num_rungs):
                    total_revenue_to_level = 0.0
                    total_qty_to_level = 0.0
                    for j in range(i + 1):
                        total_revenue_to_level += self.sell_quantities[j] * self.sell_prices[j]
                        total_qty_to_level += self.sell_quantities[j]
                    if total_qty_to_level > 0:
                        avg_sell_price = total_revenue_to_level / total_qty_to_level
                    else:
                        avg_sell_price = 0.0
                    avg_sell_prices.append(avg_sell_price)
            else:
                raise ValueError(f"Cannot normalize sell quantities: total is zero or negative")
        else:
            logger.debug(f"Volume matching verified in results: buy={total_buy_qty:.6f}, sell={total_sell_qty:.6f}, diff={qty_diff:.9f}")
        
        # Verify that highest buy average equals current price exactly
        if len(avg_buy_prices) > 0 and len(self.buy_prices) > 0:
            highest_buy_index = len(avg_buy_prices) - 1
            expected_highest_avg = self.buy_prices[highest_buy_index]  # Should equal current_price
            actual_highest_avg = avg_buy_prices[highest_buy_index]
            if abs(actual_highest_avg - expected_highest_avg) > 0.0001:
                logger.warning(f"Highest buy average mismatch: expected {expected_highest_avg:.2f}, got {actual_highest_avg:.2f}. "
                           f"Buy price: {self.buy_prices[highest_buy_index]:.2f}, Quantity: {self.buy_quantities[highest_buy_index]:.4f}")
                # Force correct value
                avg_buy_prices[highest_buy_index] = self.buy_prices[highest_buy_index]
            else:
                logger.debug(f"Highest buy average verified: {actual_highest_avg:.2f} = {expected_highest_avg:.2f}")
        
        # Verify that first sell average equals first sell price exactly
        if len(avg_sell_prices) > 0 and len(self.sell_prices) > 0:
            expected_first_avg = self.sell_prices[0]
            actual_first_avg = avg_sell_prices[0]
            if abs(actual_first_avg - expected_first_avg) > 0.0001:
                logger.warning(f"First sell average mismatch: expected {expected_first_avg:.2f}, got {actual_first_avg:.2f}. "
                           f"Sell price: {self.sell_prices[0]:.2f}, Quantity: {self.sell_quantities[0]:.4f}")
                # Force correct value
                avg_sell_prices[0] = self.sell_prices[0]
            else:
                logger.debug(f"First sell average verified: {actual_first_avg:.2f} = {expected_first_avg:.2f}")
        
        # Verify buy averages are monotonically increasing (as we go from lowest to highest price)
        # The average should increase as we move closer to current price
        for i in range(1, len(avg_buy_prices)):
            if avg_buy_prices[i] < avg_buy_prices[i-1]:
                logger.warning(f"Buy average decreases at index {i}: {avg_buy_prices[i-1]:.2f} -> {avg_buy_prices[i]:.2f}")
        
        # Verify sell averages are monotonically increasing (as we sell at higher prices)
        for i in range(1, len(avg_sell_prices)):
            if avg_sell_prices[i] < avg_sell_prices[i-1]:
                logger.warning(f"Sell average decreases at index {i}: {avg_sell_prices[i-1]:.2f} -> {avg_sell_prices[i]:.2f}")
        
        total_cost = cumulative_buy_cost[0] if cumulative_buy_cost else 0.0
        total_revenue = cumulative_sell_revenue[-1]
        total_profit = total_revenue - total_cost
        profit_pct = (total_profit / total_cost * 100) if total_cost > 0 else 0
        
        return {
            'buy_prices': self.buy_prices.tolist() if isinstance(self.buy_prices, np.ndarray) else self.buy_prices,
            'buy_quantities': self.buy_quantities,
            'sell_prices': self.sell_prices.tolist() if isinstance(self.sell_prices, np.ndarray) else self.sell_prices,
            'sell_quantities': self.sell_quantities,
            'cumulative_buy_cost': cumulative_buy_cost,
            'cumulative_buy_qty': cumulative_buy_qty,
            'cumulative_sell_revenue': cumulative_sell_revenue,
            'cumulative_sell_qty': cumulative_sell_qty,
            'avg_buy_prices': avg_buy_prices,
            'avg_sell_prices': avg_sell_prices,
            'total_cost': total_cost,
            'total_revenue': total_revenue,
            'total_profit': total_profit,
            'profit_pct': profit_pct,
            'num_rungs': self.num_rungs,
            'budget': self.budget,
            'bottom_buy_price': bottom_buy_price,
            'top_sell_price': top_sell_price,
            'ladder_shape_key': self.ladder_shape_key,
            'ladder_shape_label': self.ladder_shape_label,
            'ladder_shape_description': shape_description
        }


def generate_excel(results: Dict, filename: str):
    """Generate Excel file with order ladder and statistics."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.cell.cell import MergedCell

    wb = Workbook()
    ws = wb.active
    ws.title = "Order Ladder"
    
    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws['A1'] = "Staggered Order Ladder"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:F1')
    
    # Summary section
    row = 3
    ws[f'A{row}'] = "Summary"
    ws[f'A{row}'].font = Font(bold=True, size=14)
    row += 1
    
    buy_prices = list(results.get('buy_prices', []))
    ladder_depth_pct = compute_ladder_depth_pct(buy_prices)
    ladder_shape_summary = summarize_ladder_shape(results)
    
    summary_data = [
        ['Total Budget', f"${results['budget']:,.2f}"],
        ['Starting Price', f"${results['buy_prices'][-1]:,.2f}" if results['buy_prices'] else "$0.00"],
        ['Number of Rungs', results['num_rungs']],
        ['Ladder Shape', ladder_shape_summary],
        ['Ladder Depth', f"{ladder_depth_pct:.2f}%"],
        ['Bottom Buy Price', f"${results['bottom_buy_price']:,.2f}"],
        ['Top Sell Price', f"${results['top_sell_price']:,.2f}"],
        ['Total Cost', f"${results['total_cost']:,.2f}"],
        ['Total Revenue', f"${results['total_revenue']:,.2f}"],
        ['Total Profit', f"${results['total_profit']:,.2f}"],
        ['Profit Percentage', f"{results['profit_pct']:.2f}%"]
    ]
    
    for label, value in summary_data:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
    
    row += 1
    
    # Buy orders header
    ws[f'A{row}'] = "Buy Orders"
    ws[f'A{row}'].font = Font(bold=True, size=14)
    row += 1
    
    headers = ['Buy Price', 'Quantity', 'Cost', 'Cumulative Cost', 'Avg Buy Price']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    row += 1
    
    # Buy orders data (reversed: order 1 = highest price, order 10 = lowest price)
    buy_prices_reversed = list(reversed(results['buy_prices']))
    buy_quantities_reversed = list(reversed(results['buy_quantities']))
    cumulative_buy_cost_reversed = list(reversed(results['cumulative_buy_cost']))
    avg_buy_prices_reversed = list(reversed(results['avg_buy_prices']))
    
    for i in range(results['num_rungs']):
        ws.cell(row=row, column=1).value = buy_prices_reversed[i]
        ws.cell(row=row, column=2).value = buy_quantities_reversed[i]
        ws.cell(row=row, column=3).value = buy_prices_reversed[i] * buy_quantities_reversed[i]
        ws.cell(row=row, column=4).value = cumulative_buy_cost_reversed[i]
        ws.cell(row=row, column=5).value = avg_buy_prices_reversed[i]
        
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = border
            ws.cell(row=row, column=col).number_format = '#,##0.00'
        
        row += 1
    
    row += 1
    
    # Sell orders header
    ws[f'A{row}'] = "Sell Orders"
    ws[f'A{row}'].font = Font(bold=True, size=14)
    row += 1
    
    headers = ['Sell Price', 'Quantity', 'Revenue', 'Cumulative Revenue', 'Avg Sell Price']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center')
    
    row += 1
    
    # Sell orders data
    for i in range(results['num_rungs']):
        ws.cell(row=row, column=1).value = results['sell_prices'][i]
        ws.cell(row=row, column=2).value = results['sell_quantities'][i]
        ws.cell(row=row, column=3).value = results['sell_prices'][i] * results['sell_quantities'][i]
        ws.cell(row=row, column=4).value = results['cumulative_sell_revenue'][i]
        ws.cell(row=row, column=5).value = results['avg_sell_prices'][i]
        
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = border
            ws.cell(row=row, column=col).number_format = '#,##0.00'
        
        row += 1
    
    # Auto-adjust column widths
    for col_idx, column in enumerate(ws.columns, start=1):
        max_length = 0
        # Skip if first cell is a MergedCell
        if isinstance(column[0], MergedCell):
            continue
        
        column_letter = get_column_letter(col_idx)
        for cell in column:
            try:
                if not isinstance(cell, MergedCell) and cell.value is not None:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
            except:
                pass
        if max_length > 0:
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(filename)
    logger.info(f"Excel file saved: {filename}")


@dataclass(frozen=True)
class OrderSideConfig:
    side: str
    order_label: str
    price_key: str
    quantity_key: str
    cumulative_key: str
    average_key: str
    table_title: str
    chart_title: str
    individual_label: str
    average_label: str
    xlabel: str
    column_labels: List[str]
    price_line_color: str
    average_line_color: str
    price_axis_color: str
    price_annotation_color: str
    quantity_line_color: str
    quantity_area_color: str
    quantity_axis_color: str
    quantity_annotation_color: str
    reverse_for_display: bool = False


def _prepare_order_sequences(results: Dict, config: OrderSideConfig) -> Tuple[List[float], List[float], List[float], List[float]]:
    prices = list(results.get(config.price_key, []))
    quantities = list(results.get(config.quantity_key, []))
    cumulative = list(results.get(config.cumulative_key, []))
    averages = list(results.get(config.average_key, []))

    if config.reverse_for_display:
        prices.reverse()
        quantities.reverse()
        cumulative.reverse()
        averages.reverse()

    return prices, quantities, cumulative, averages


def _build_table_rows(prices: List[float], quantities: List[float], cumulative: List[float], averages: List[float]) -> List[List[str]]:
    rows: List[List[str]] = []
    for price, qty, cum, avg in zip(prices, quantities, cumulative, averages):
        rows.append([
            f"${price:,.2f}",
            f"{qty:,.4f}",
            f"${price * qty:,.2f}",
            f"${cum:,.2f}",
            f"${avg:,.2f}"
        ])
    return rows


def _create_summary_figure(results: Dict, configs: Tuple[OrderSideConfig, OrderSideConfig]):
    import matplotlib.pyplot as plt

    fig = plt.figure(figsize=(11, 8.5))
    fig.text(
        0.5,
        0.95,
        'Staggered Order Ladder',
        ha='center',
        fontsize=16,
        fontweight='bold'
    )

    buy_prices = list(results.get('buy_prices', []))
    ladder_depth_pct = compute_ladder_depth_pct(buy_prices)
    ladder_shape_summary = summarize_ladder_shape(results)
    
    summary_text = (
        f"\nSummary:\n"
        f"  Total Budget: ${results['budget']:,.2f}\n"
        f"  Number of Rungs: {results['num_rungs']}\n"
        f"  Ladder Shape: {ladder_shape_summary}\n"
        f"  Ladder Depth: {ladder_depth_pct:.2f}%\n"
        f"  Bottom Buy Price: ${results['bottom_buy_price']:,.2f}\n"
        f"  Top Sell Price: ${results['top_sell_price']:,.2f}\n"
        f"  Total Cost: ${results['total_cost']:,.2f}\n"
        f"  Total Revenue: ${results['total_revenue']:,.2f}\n"
        f"  Total Profit: ${results['total_profit']:,.2f}\n"
        f"  Profit Percentage: {results['profit_pct']:.2f}%\n"
    )

    fig.text(
        0.1,
        0.85,
        summary_text,
        fontsize=10,
        family='monospace',
        verticalalignment='top'
    )

    for index, config in enumerate(configs, start=1):
        ax = fig.add_subplot(2, 1, index)
        ax.axis('tight')
        ax.axis('off')

        prices, quantities, cumulative, averages = _prepare_order_sequences(results, config)
        table_data = _build_table_rows(prices, quantities, cumulative, averages)

        table = ax.table(
            cellText=table_data,
            colLabels=config.column_labels,
            cellLoc='center',
            loc='center'
        )
        table.auto_set_font_size(False)
        table.set_fontsize(8)
        table.scale(1, 1.5)
        ax.set_title(config.table_title, fontweight='bold', pad=20)

    plt.tight_layout()
    return fig


def _create_order_graph_figure(results: Dict, config: OrderSideConfig):
    import matplotlib.pyplot as plt

    prices, quantities, _, averages = _prepare_order_sequences(results, config)
    fig, ax_price = plt.subplots(figsize=(11, 8.5))

    if not prices:
        ax_price.text(
            0.5,
            0.5,
            f"No {config.side} order data available.",
            transform=ax_price.transAxes,
            ha='center',
            va='center',
            fontsize=14,
            fontweight='bold'
        )
        ax_price.axis('off')
        return fig

    x_positions = list(range(len(prices)))
    ax_qty = ax_price.twinx()

    ax_qty.fill_between(
        x_positions,
        [0] * len(quantities),
        quantities,
        alpha=0.2,
        color=config.quantity_area_color,
        label='Quantity',
        step='mid'
    )
    ax_qty.plot(
        x_positions,
        quantities,
        marker='^',
        markersize=5,
        linewidth=1.5,
        color=config.quantity_line_color,
        linestyle=':',
        alpha=0.7
    )
    ax_qty.set_ylabel('Quantity', fontsize=12, fontweight='bold', color=config.quantity_axis_color)
    ax_qty.tick_params(axis='y', labelcolor=config.quantity_axis_color)

    ax_price.plot(
        x_positions,
        prices,
        marker='o',
        linewidth=2.5,
        markersize=9,
        color=config.price_line_color,
        label=config.individual_label,
        linestyle='-',
        zorder=3
    )
    ax_price.plot(
        x_positions,
        averages,
        marker='s',
        linewidth=2.5,
        markersize=7,
        color=config.average_line_color,
        label=config.average_label,
        linestyle='--',
        zorder=3
    )
    ax_price.set_ylabel('Price ($)', fontsize=12, fontweight='bold', color=config.price_axis_color)
    ax_price.tick_params(axis='y', labelcolor=config.price_axis_color)
    ax_price.set_title(config.chart_title, fontsize=14, fontweight='bold')
    ax_price.grid(True, alpha=0.3, axis='both', linestyle='--', linewidth=0.5)

    for idx in range(0, len(x_positions), 2):
        ax_price.annotate(
            f'${prices[idx]:.0f}',
            xy=(x_positions[idx], prices[idx]),
            xytext=(0, 10),
            textcoords='offset points',
            fontsize=8,
            ha='center',
            color=config.price_annotation_color,
            alpha=0.7
        )

    top_quantity_indices = sorted(
        range(len(quantities)),
        key=lambda i: quantities[i],
        reverse=True
    )[:3]
    for idx in top_quantity_indices:
        ax_qty.annotate(
            f'{quantities[idx]:.1f}',
            xy=(x_positions[idx], quantities[idx]),
            xytext=(0, 5),
            textcoords='offset points',
            fontsize=8,
            ha='center',
            color=config.quantity_annotation_color,
            fontweight='bold',
            alpha=0.8
        )

    ax_price.set_xticks(x_positions)
    ax_price.set_xticklabels(
        [f'{config.order_label} {idx + 1}' for idx in range(len(prices))],
        rotation=0,
        ha='center',
        fontsize=10
    )
    ax_price.set_xlabel(config.xlabel, fontsize=12, fontweight='bold')

    lines_price, labels_price = ax_price.get_legend_handles_labels()
    lines_qty, labels_qty = ax_qty.get_legend_handles_labels()
    ax_price.legend(
        lines_price + lines_qty,
        labels_price + labels_qty,
        loc='upper left',
        fontsize=10,
        framealpha=0.9,
        edgecolor='gray',
        fancybox=True
    )

    plt.tight_layout()
    return fig


def _calculate_point_sizes(buy_quantities: List[float], sell_quantities: List[float]) -> Tuple[List[float], List[float], float, float]:
    all_quantities = buy_quantities + sell_quantities
    if not all_quantities:
        return [], [], 0.0, 0.0

    min_qty = min(all_quantities)
    max_qty = max(all_quantities)
    qty_range = max_qty - min_qty if max_qty > min_qty else 1

    min_point_size = 50
    max_point_size = 500

    def scale_size(qty: float) -> float:
        if qty_range == 0:
            return (min_point_size + max_point_size) / 2
        normalized = (qty - min_qty) / qty_range
        return min_point_size + normalized * (max_point_size - min_point_size)

    buy_point_sizes = [scale_size(qty) for qty in buy_quantities]
    sell_point_sizes = [scale_size(qty) for qty in sell_quantities]

    return buy_point_sizes, sell_point_sizes, min_qty, max_qty


def _calculate_bar_width(positions: List[float]) -> float:
    if not positions:
        return 1
    if len(positions) > 1:
        return (positions[1] - positions[0]) * 0.8
    return abs(positions[0]) * 0.1 if positions[0] != 0 else 1


def _calculate_volume_bars(y_min_price: float, y_range_price: float, buy_volumes: List[float], sell_volumes: List[float]) -> Tuple[float, List[float], List[float]]:
    all_volumes = buy_volumes + sell_volumes
    max_volume = max(all_volumes) if all_volumes else 1
    volume_bar_height = y_range_price * 0.15 if y_range_price else 1
    volume_scale = volume_bar_height / max_volume if max_volume > 0 else 1

    volume_bottom = y_min_price - y_range_price * 0.05 if y_range_price else y_min_price - 0.5
    buy_volume_heights = [vol * volume_scale for vol in buy_volumes]
    sell_volume_heights = [vol * volume_scale for vol in sell_volumes]

    return volume_bottom, buy_volume_heights, sell_volume_heights


def _create_combined_ladder_figure(results: Dict):
    import matplotlib.pyplot as plt

    fig, ax = plt.subplots(figsize=(11, 8.5))

    buy_prices = list(results.get('buy_prices', []))
    buy_quantities = list(results.get('buy_quantities', []))
    avg_buy_prices = list(results.get('avg_buy_prices', []))

    sell_prices = list(results.get('sell_prices', []))
    sell_quantities = list(results.get('sell_quantities', []))
    avg_sell_prices = list(results.get('avg_sell_prices', []))

    if buy_prices and avg_buy_prices:
        highest_index = len(buy_prices) - 1
        if abs(avg_buy_prices[highest_index] - buy_prices[highest_index]) > 0.01:
            logger.warning(
                "Graph: Highest buy average %.2f != highest buy price %.2f",
                avg_buy_prices[highest_index],
                buy_prices[highest_index]
            )

    if sell_prices and avg_sell_prices:
        if abs(avg_sell_prices[0] - sell_prices[0]) > 0.01:
            logger.warning(
                "Graph: First sell average %.2f != first sell price %.2f, correcting...",
                avg_sell_prices[0],
                sell_prices[0]
            )
            avg_sell_prices = list(avg_sell_prices)
            avg_sell_prices[0] = sell_prices[0]

    if not (buy_prices and sell_prices):
        ax.text(
            0.5,
            0.5,
            "Insufficient order data for combined ladder chart.",
            transform=ax.transAxes,
            ha='center',
            va='center',
            fontsize=14,
            fontweight='bold'
        )
        ax.axis('off')
        return fig

    buy_point_sizes, sell_point_sizes, min_qty, max_qty = _calculate_point_sizes(buy_quantities, sell_quantities)

    min_buy_price = min(buy_prices)
    max_buy_price = max(buy_prices)
    min_sell_price = min(sell_prices)
    max_sell_price = max(sell_prices)

    price_midpoint = (max_buy_price + min_sell_price) / 2

    buy_x_positions = [price - price_midpoint for price in buy_prices]
    sell_x_positions = [price - price_midpoint for price in sell_prices]

    buy_volumes = [qty * price for qty, price in zip(buy_quantities, buy_prices)]
    sell_volumes = [qty * price for qty, price in zip(sell_quantities, sell_prices)]

    y_min_price = min(min(avg_buy_prices), min(avg_sell_prices))
    y_max_price = max(max(avg_buy_prices), max(avg_sell_prices))
    y_range_price = y_max_price - y_min_price

    buy_bar_width = _calculate_bar_width(buy_x_positions)
    sell_bar_width = _calculate_bar_width(sell_x_positions)

    volume_bottom, buy_volume_heights, sell_volume_heights = _calculate_volume_bars(
        y_min_price,
        y_range_price,
        buy_volumes,
        sell_volumes
    )

    ax.bar(
        buy_x_positions,
        buy_volume_heights,
        width=buy_bar_width,
        bottom=volume_bottom,
        alpha=0.5,
        color='#81C784',
        edgecolor='#2E7D32',
        linewidth=1,
        label='Buy Volume',
        zorder=2
    )

    ax.bar(
        sell_x_positions,
        sell_volume_heights,
        width=sell_bar_width,
        bottom=volume_bottom,
        alpha=0.5,
        color='#EF5350',
        edgecolor='#C62828',
        linewidth=1,
        label='Sell Volume',
        zorder=2
    )

    ax.scatter(
        buy_x_positions,
        avg_buy_prices,
        s=buy_point_sizes,
        alpha=0.6,
        color='#2E7D32',
        edgecolors='#1B5E20',
        linewidths=1.5,
        label='Buy Orders',
        zorder=3
    )

    ax.scatter(
        sell_x_positions,
        avg_sell_prices,
        s=sell_point_sizes,
        alpha=0.6,
        color='#C62828',
        edgecolors='#B71C1C',
        linewidths=1.5,
        label='Sell Orders',
        zorder=3
    )

    ax.axvline(x=0, color='gray', linestyle='--', linewidth=2, alpha=0.7, zorder=1)

    all_x_positions = buy_x_positions + sell_x_positions
    x_min = min(all_x_positions)
    x_max = max(all_x_positions)
    x_range = x_max - x_min
    x_padding = x_range * 0.05
    ax.set_xlim(x_min - x_padding, x_max + x_padding)

    max_volume_bar_height = max(
        max(buy_volume_heights) if buy_volume_heights else 0,
        max(sell_volume_heights) if sell_volume_heights else 0
    )
    volume_bar_top = volume_bottom + max_volume_bar_height
    y_min_with_volume = min(volume_bottom, y_min_price) - y_range_price * 0.02
    y_max_with_volume = max(volume_bar_top, y_max_price) + y_range_price * 0.02
    ax.set_ylim(y_min_with_volume, y_max_with_volume)

    x_ticks = buy_x_positions + sell_x_positions
    x_labels = [f'${price:.2f}' for price in buy_prices + sell_prices]
    ax.set_xticks(x_ticks)
    ax.set_xticklabels(x_labels, rotation=45, ha='right', fontsize=8)

    ax.set_xlabel('Order Price ($)', fontsize=12, fontweight='bold')
    ax.set_ylabel('Average Price ($)', fontsize=12, fontweight='bold')
    ax.set_title(
        'Combined Buy/Sell Ladder: Average Price vs Order Price\n(Point size = Order Quantity)',
        fontsize=14,
        fontweight='bold',
        pad=20
    )

    ax.grid(True, alpha=0.3, linestyle='--', linewidth=0.5, zorder=0)
    ax.legend(loc='upper left', fontsize=10, framealpha=0.9, edgecolor='gray', fancybox=True)

    min_qty_formatted = f"{min_qty:.2f}" if min_qty < 1000 else f"{min_qty:.0f}"
    max_qty_formatted = f"{max_qty:.2f}" if max_qty < 1000 else f"{max_qty:.0f}"
    size_info = f"Point size scale:\nMin: {min_qty_formatted}\nMax: {max_qty_formatted}"
    ax.text(
        0.02,
        0.98,
        size_info,
        transform=ax.transAxes,
        fontsize=9,
        verticalalignment='top',
        bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.5)
    )

    ax.text(
        0.02,
        0.02,
        'BUY ORDERS',
        transform=ax.transAxes,
        fontsize=11,
        fontweight='bold',
        color='#2E7D32',
        bbox=dict(boxstyle='round', facecolor='lightgreen', alpha=0.7)
    )
    ax.text(
        0.98,
        0.02,
        'SELL ORDERS',
        transform=ax.transAxes,
        fontsize=11,
        fontweight='bold',
        color='#C62828',
        bbox=dict(boxstyle='round', facecolor='lightcoral', alpha=0.7),
        ha='right'
    )

    plt.tight_layout()
    return fig


def generate_pdf(results: Dict, filename: str):
    """Generate PDF file with order ladder, statistics, and graphs."""
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_pdf import PdfPages

    buy_config = OrderSideConfig(
        side='buy',
        order_label='Buy order',
        price_key='buy_prices',
        quantity_key='buy_quantities',
        cumulative_key='cumulative_buy_cost',
        average_key='avg_buy_prices',
        table_title='Buy Orders',
        chart_title='Buy Ladder: Price and Quantity by Order Rung',
        individual_label='Individual Buy Price',
        average_label='Average Buy Price',
        xlabel='Buy Order',
        column_labels=['Buy Price', 'Quantity', 'Cost', 'Cumulative Cost', 'Avg Buy Price'],
        price_line_color='#1B5E20',
        average_line_color='#2E7D32',
        price_axis_color='#2E7D32',
        price_annotation_color='#1B5E20',
        quantity_line_color='#66BB6A',
        quantity_area_color='#81C784',
        quantity_axis_color='#66BB6A',
        quantity_annotation_color='#66BB6A',
        reverse_for_display=True
    )

    sell_config = OrderSideConfig(
        side='sell',
        order_label='Sell order',
        price_key='sell_prices',
        quantity_key='sell_quantities',
        cumulative_key='cumulative_sell_revenue',
        average_key='avg_sell_prices',
        table_title='Sell Orders',
        chart_title='Sell Ladder: Price and Quantity by Order Rung',
        individual_label='Individual Sell Price',
        average_label='Average Sell Price',
        xlabel='Sell Order',
        column_labels=['Sell Price', 'Quantity', 'Revenue', 'Cumulative Revenue', 'Avg Sell Price'],
        price_line_color='#B71C1C',
        average_line_color='#C62828',
        price_axis_color='#C62828',
        price_annotation_color='#B71C1C',
        quantity_line_color='#E53935',
        quantity_area_color='#EF5350',
        quantity_axis_color='#E53935',
        quantity_annotation_color='#E53935'
    )

    with PdfPages(filename) as pdf:
        summary_fig = _create_summary_figure(results, (buy_config, sell_config))
        pdf.savefig(summary_fig, bbox_inches='tight')
        plt.close(summary_fig)

        for config in (buy_config, sell_config):
            fig = _create_order_graph_figure(results, config)
            pdf.savefig(fig, bbox_inches='tight')
            plt.close(fig)

        combined_fig = _create_combined_ladder_figure(results)
        pdf.savefig(combined_fig, bbox_inches='tight')
        plt.close(combined_fig)
    
    logger.info(f"PDF file saved: {filename}")


def get_smart_input(args: Optional[argparse.Namespace] = None) -> Dict:
    """
    Get input for ladder calculation with intelligent defaults.
    
    Args:
        args: Optional argparse.Namespace with command-line arguments
    
    Returns:
        Dictionary with calculation results
    """
    print("\n" + "="*60)
    print("Staggered Order Ladder Calculator")
    print("="*60 + "\n")
    
    argv_tokens = sys.argv[1:] if args else []
    
    def argument_was_provided(flag_names: Tuple[str, ...]) -> bool:
        if not argv_tokens:
            return False
        for token in argv_tokens:
            for flag in flag_names:
                if token == flag or token.startswith(f"{flag}="):
                    return True
        return False
    
    def prompt_depth_value(default_value: float) -> float:
        while True:
            prompt = (
                f"How deep do you want to go from the current price (%) "
                f"[default {default_value:.2f}%]: "
            )
            entry = input(prompt).strip()
            if not entry:
                depth_val = default_value
                print(f"Using ladder depth: {depth_val:.2f}%")
                return depth_val
            try:
                depth_val = float(entry)
            except ValueError:
                print("Please enter a valid number.")
                continue
            if depth_val <= 0 or depth_val >= 100:
                print("Depth must be greater than 0% and less than 100%.")
                continue
            return depth_val
    
    def prompt_profit_target_value(default_value: float) -> float:
        while True:
            entry = input(
                f"Enter profit target percentage [default {default_value:.2f}%]: "
            ).strip()
            if not entry:
                profit_val = default_value
                print(f"Using profit target: {profit_val:.2f}%")
                return profit_val
            try:
                profit_val = float(entry)
            except ValueError:
                print("Please enter a valid number.")
                continue
            if profit_val < 10 or profit_val > 200:
                print("Profit target must be between 10% and 200%.")
                continue
            return profit_val
    
    def prompt_ladder_shape(default_key: str) -> str:
        shape_keys = list(LADDER_SHAPE_PROFILES.keys())
        print("\nLadder shape options:")
        for idx, key in enumerate(shape_keys, start=1):
            profile = LADDER_SHAPE_PROFILES[key]
            print(f"  {idx}. {profile.label} ({key}) - {profile.description}")
        default_index = shape_keys.index(default_key) + 1 if default_key in shape_keys else 1
        while True:
            selection = input(
                f"Select ladder shape [default {default_index}]: "
            ).strip().lower()
            if not selection:
                chosen_key = default_key
            elif selection.isdigit():
                option_idx = int(selection)
                if 1 <= option_idx <= len(shape_keys):
                    chosen_key = shape_keys[option_idx - 1]
                else:
                    print("Please choose a valid option number.")
                    continue
            else:
                try:
                    chosen_key = normalize_ladder_shape_key(selection)
                except ValueError:
                    print("Invalid shape. Please choose one of the listed options by number or name.")
                    continue
            profile = LADDER_SHAPE_PROFILES[chosen_key]
            print(f"Selected ladder shape: {profile.label} - {profile.description}")
            return chosen_key
    
    # Get budget (required)
    if args and args.budget:
        budget = args.budget
        print(f"Budget: ${budget:,.2f}")
    else:
        while True:
            try:
                budget_str = input("Enter total budget for buy orders: $").strip()
                if not budget_str:
                    print("Budget is required.")
                    continue
                budget = float(budget_str)
                if budget <= 0:
                    print("Budget must be positive.")
                    continue
                break
            except ValueError:
                print("Please enter a valid number.")
    
    # Get ladder depth before other market inputs
    default_depth = args.price_range if (args and args.price_range is not None) else 30.0
    if args and argument_was_provided(('--price-range',)):
        buy_price_range_pct = float(args.price_range)
        print(f"Ladder depth: {buy_price_range_pct:.2f}%")
    else:
        buy_price_range_pct = prompt_depth_value(default_depth)
    
    if buy_price_range_pct <= 0 or buy_price_range_pct >= 100:
        raise ValueError("Buy price range percentage must be greater than 0 and less than 100.")
    
    # Ask user for ladder shape preference with descriptors
    default_shape_key = "deep-dive"
    if args and getattr(args, "ladder_shape", None):
        try:
            ladder_shape_key = normalize_ladder_shape_key(args.ladder_shape)
        except ValueError as exc:
            raise ValueError(str(exc))
        profile = LADDER_SHAPE_PROFILES[ladder_shape_key]
        print(f"Ladder shape: {profile.label} - {profile.description}")
    else:
        ladder_shape_key = prompt_ladder_shape(default_shape_key)
    
    # Get current price (required for profit target method)
    if args and args.price:
        current_price = args.price
        print(f"Current Price: ${current_price:,.2f}")
    else:
        while True:
            try:
                price_str = input("Enter current price: $").strip()
                if not price_str:
                    print("Price is required.")
                    continue
                current_price = float(price_str)
                if current_price <= 0:
                    print("Price must be positive.")
                    continue
                break
            except ValueError:
                print("Please enter a valid number.")
    
    # Get profit target with confirmation support
    default_profit_target = args.profit_target if (args and args.profit_target is not None) else 75.0
    if args and argument_was_provided(('--profit-target',)):
        profit_target = float(args.profit_target)
        print(f"Profit Target: {profit_target:.2f}%")
    else:
        profit_target = prompt_profit_target_value(default_profit_target)
    
    if profit_target < 10 or profit_target > 200:
        raise ValueError("Profit target must be between 10% and 200%.")
    
    # Get number of orders (with default 10)
    default_num_rungs = args.num_rungs if (args and args.num_rungs is not None) else 10
    if args and argument_was_provided(('--num-rungs',)):
        num_rungs = int(args.num_rungs)
        if num_rungs < 1:
            raise ValueError("Number of orders must be at least 1.")
        print(f"Number of Orders: {num_rungs}")
    else:
        while True:
            try:
                orders_str = input(
                    f"Enter number of orders to place (default {default_num_rungs}): "
                ).strip()
                if not orders_str:
                    num_rungs = default_num_rungs
                    print(f"Using default number of orders: {num_rungs}")
                    break
                num_rungs = int(orders_str)
                if num_rungs < 1:
                    print("Number of orders must be at least 1.")
                    continue
                break
            except ValueError:
                print("Please enter a valid integer.")
    
    calculator = StaggeredLadderCalculator(budget, num_rungs)
    can_prompt_user = hasattr(sys.stdin, "isatty") and sys.stdin.isatty()
    results: Dict = {}
    
    def print_configuration():
        profile = LADDER_SHAPE_PROFILES[ladder_shape_key]
        print("\nConfiguration:")
        print(f"  Total budget: ${budget:,.2f}")
        print(f"  Number of orders: {num_rungs}")
        print(f"  Profit target: {profit_target:.2f}%")
        print(f"  Ladder depth: {buy_price_range_pct:.2f}% below current price")
        print(f"  Ladder shape: {profile.label} - {profile.description}")
        print(f"  Profit mode: overall")
    
    while True:
        print_configuration()
        results = calculator.calculate_from_profit_target(
            current_price,
            profit_target,
            buy_price_range_pct,
            ladder_shape_key=ladder_shape_key
        )
        
        bottom_buy_price = results['bottom_buy_price']
        top_sell_price = results['top_sell_price']
        
        print("\nLadder depth confirmation:")
        print(f"  Bottom buy price: ${bottom_buy_price:,.2f}")
        print(f"  Top sell price:   ${top_sell_price:,.2f}")
        
        if not can_prompt_user:
            break
        
        confirmation = input("Confirm these ladder bounds? [Y/n]: ").strip().lower()
        if confirmation in ("", "y", "yes"):
            break
        
        print("\nLet's adjust the ladder depth.")
        buy_price_range_pct = prompt_depth_value(buy_price_range_pct)
        adjust_profit = input("Adjust profit target as well? [y/N]: ").strip().lower()
        if adjust_profit in ("y", "yes"):
            profit_target = prompt_profit_target_value(profit_target)
    
    return results


def main():
    """Main function with command-line argument support."""
    parser = argparse.ArgumentParser(
        description='Staggered Order Ladder Calculator - Configurable ladder shapes with overall profit mode',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Minimal input (only budget and price required)
  python staggered_ladder.py --budget 10000 --price 50
  
  # Customize profit target and number of orders
  python staggered_ladder.py --budget 10000 --price 50 --profit-target 75 --num-rungs 15

Defaults:
  - Profit target: 75%
  - Number of orders: 10
  - Ladder depth: 30% below current price
  - Ladder shape: Deep Dive (heavier lower buys, heavier higher sells)
  - Profit mode: overall
        """
    )
    
    parser.add_argument(
        '--budget',
        type=float,
        help='Total budget for buy orders (required if not using interactive mode)'
    )
    parser.add_argument(
        '--price',
        type=float,
        help='Current price (required if not using interactive mode)'
    )
    parser.add_argument(
        '--profit-target',
        type=float,
        default=75.0,
        help='Target profit percentage (default: 75%%)'
    )
    parser.add_argument(
        '--num-rungs',
        type=int,
        default=10,
        help='Number of orders to place (default: 10)'
    )
    parser.add_argument(
        '--price-range',
        type=float,
        default=30.0,
        help='Ladder depth percentage below current price (default: 30%%)'
    )
    parser.add_argument(
        '--ladder-shape',
        type=str,
        help='Ladder shape profile (options: balanced, deep-dive, springboard, glide-path)'
    )
    parser.add_argument(
        '--output-prefix',
        type=str,
        default=None,
        help='Optional prefix for output filenames'
    )
    
    args = parser.parse_args()
    
    try:
        results = get_smart_input(args)
        
        # Create output directory relative to script location
        script_dir = os.path.dirname(os.path.abspath(__file__))
        output_dir = os.path.join(script_dir, "output")
        os.makedirs(output_dir, exist_ok=True)
        
        # Generate unique filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        prefix = f"{args.output_prefix}_" if args.output_prefix else ""
        excel_filename = os.path.join(output_dir, f"{prefix}staggered_ladder_{timestamp}.xlsx")
        pdf_filename = os.path.join(output_dir, f"{prefix}staggered_ladder_{timestamp}.pdf")
        
        # Generate outputs
        print("\nGenerating outputs...")
        generate_excel(results, excel_filename)
        generate_pdf(results, pdf_filename)
        
        print(f"\n[OK] Excel file generated: {excel_filename}")
        print(f"[OK] PDF file generated: {pdf_filename}")
        
        # Calculate total volumes (quantities)
        total_buy_volume = sum(results['buy_quantities'])
        total_sell_volume = sum(results['sell_quantities'])
        
        ladder_depth_pct = compute_ladder_depth_pct(list(results.get('buy_prices', [])))
        ladder_shape_summary = summarize_ladder_shape(results)
        
        print(f"\nSummary:")
        print(f"  Ladder Shape: {ladder_shape_summary}")
        print(f"  Ladder Depth: {ladder_depth_pct:.2f}%")
        print(f"  Bottom Buy Price: ${results['bottom_buy_price']:,.2f}")
        print(f"  Top Sell Price: ${results['top_sell_price']:,.2f}")
        print(f"  Total Cost: ${results['total_cost']:,.2f}")
        print(f"  Total Revenue: ${results['total_revenue']:,.2f}")
        print(f"  Total Profit: ${results['total_profit']:,.2f}")
        print(f"  Profit Percentage: {results['profit_pct']:.2f}%")
        print(f"  Total Buy Volume: {total_buy_volume:,.4f}")
        print(f"  Total Sell Volume: {total_sell_volume:,.4f}")
        
    except KeyboardInterrupt:
        print("\n\nOperation cancelled by user.")
        sys.exit(0)
    except Exception as e:
        logger.error(f"Error: {str(e)}", exc_info=True)
        print(f"\nError: {str(e)}")
        sys.exit(1)


if __name__ == '__main__':
    main()

